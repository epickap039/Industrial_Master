import socket
import uvicorn
import pyodbc
import pandas as pd
import openpyxl
from fastapi import FastAPI, HTTPException, Request, Response, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fastapi.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
from typing import Dict, Any, List, Optional
from pydantic import BaseModel, ConfigDict
import io
import os
import re
import uuid

# 1. CONFIGURACIÓN SQL (Auto-Detectada con Driver 18 Prioritario)
DB_SERVER = '192.168.1.73'
DB_PORT = 1433
DB_DATABASE = 'DB_Materiales_Industrial'

# Detectar el mejor driver disponible (18 > 17 > otros)
try:
    available_drivers = [d for d in pyodbc.drivers() if 'SQL Server' in d]
    if available_drivers:
        # Tomar el último (usualmente la versión más reciente, ej: ODBC Driver 18)
        best_driver = available_drivers[-1]
        DB_DRIVER = f'{{{best_driver}}}'
        print(f"SQL DRIVER SELECCIONADO: {DB_DRIVER}")
    else:
        DB_DRIVER = '{ODBC Driver 17 for SQL Server}'
        print("AVISO: No se detectaron drivers SQL. Usando default 17.")
except Exception as e:
    DB_DRIVER = '{ODBC Driver 17 for SQL Server}'
    print(f"Error detectando drivers: {e}")

# Connection String con TrustServerCertificate para Driver 18+
CONNECTION_STRING = (
    f'DRIVER={DB_DRIVER};'
    f'SERVER={DB_SERVER},{DB_PORT};'
    f'DATABASE={DB_DATABASE};'
    'Trusted_Connection=yes;'
    'TrustServerCertificate=yes;' # Crucial para Driver 18
)

# 2. SEGURIDAD Y PERMISOS
ADMIN_HOSTNAME = socket.gethostname()
try:
    ADMIN_IP = socket.gethostbyname(ADMIN_HOSTNAME)
except:
    ADMIN_IP = "127.0.0.1"

# 4. INFRAESTRUCTURA (Lifespan)
@asynccontextmanager
async def lifespan(app: FastAPI):
    print(f"--- SERVER STARTED on {ADMIN_HOSTNAME} ---")
    print(f"--- LISTENING ON 0.0.0.0:8001 ---")
    
    # Inicializaciones Seguras
    iniciar_auditoria()
    
    try:
        init_auth_db()
    except Exception as e:
        print(f"ERROR INITIALIZING AUTH DB: {e}")
    
    yield
    print("--- SERVER SHUTTING DOWN ---")

# Variables Globales de Configuración (RE-APPLIED)
REGLA_ESPEJO_ACTIVA = True

class MirrorConfig(BaseModel):
    activa: bool

app = FastAPI(title="Industrial Manager API v60.0", version="60.0", lifespan=lifespan)

# === MATERIALES APROBADOS ===
class MaterialPayload(BaseModel):
    material: str

class LoginRequest(BaseModel):
    username: str
    password: str

@app.get("/")
def read_root():
    return {"status": "online", "message": "Servidor Industrial Manager Activo"}

@app.get("/api/config/materiales")
def get_materiales():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT Material FROM Tbl_Materiales_Aprobados ORDER BY Material")
    rows = cursor.fetchall()
    conn.close()
    return [row[0] for row in rows]

@app.post("/api/config/materiales")
def add_material(payload: MaterialPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Tbl_Materiales_Aprobados (Material) VALUES (?)", (payload.material.upper(),))
        conn.commit()
        return {"mensaje": "Material agregado correctamente"}
    except pyodbc.IntegrityError:
        raise HTTPException(status_code=400, detail="El material ya existe")
    finally:
        conn.close()

@app.delete("/api/config/materiales/{material_name}")
def delete_material(material_name: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_Materiales_Aprobados WHERE Material = ?", (material_name,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Material no encontrado")
        conn.commit()
        return {"mensaje": "Material eliminado correctamente"}
    finally:
        conn.close()

class MaterialOficial(BaseModel):
    descripcion: str

@app.post("/api/materiales/oficial")
def agregar_material_oficial(payload: MaterialOficial):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO Tbl_Materiales_Aprobados (Material) VALUES (?)",
            (payload.descripcion.upper(),)
        )
        conn.commit()
        return {"status": "success", "message": "Material oficial guardado correctamente"}
    except pyodbc.IntegrityError:
        conn.rollback()
        # En caso de que el material ya exista (UNIQUE constraint)
        return {"status": "success", "message": "Material ya existe o se guardó correctamente"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en SQL Server al guardar material: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/materiales/oficial/{identificador}")
def eliminar_material_oficial(identificador: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_Materiales_Aprobados WHERE Material = ?", (identificador,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Material no encontrado")
        conn.commit()
        return {"status": "success", "message": "Material oficial eliminado correctamente"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en SQL Server al eliminar material: {str(e)}")
    finally:
        conn.close()

# === MODULO: JERARQUIA DE PROYECTOS ===
class TractoPayload(BaseModel):
    nombre: str

class TipoProyectoPayload(BaseModel):
    id_tracto: int
    nombre: str

class VersionPayload(BaseModel):
    id_tipo: int
    nombre: str

class ClientePayload(BaseModel):
    id_version: int
    nombre: str

# === MODELOS BOM ===
class RevisionPayload(BaseModel):
    nombre_revision: str

class VINPayload(BaseModel):
    vin: str
    notas: str = None

class ClonarPayload(BaseModel):
    id_revision_origen: int
    id_revision_destino: int
    rama_cliente: bool = False
    id_cliente_destino: Optional[int] = None

class PropagarPayload(BaseModel):
    codigo_pieza: str
    nueva_cantidad: float
    id_revisiones: list[int]

class EstacionPayload(BaseModel):
    id_revision: int
    nombre: str

class EnsamblePayload(BaseModel):
    id_estacion: int
    nombre: str

class BOMPayload(BaseModel):
    id_ensamble: int
    codigo_pieza: str
    cantidad: float

class AsignarRevisionPayload(BaseModel):
    id_revision_asignada: Optional[int] = None

class LogAuditoriaPayload(BaseModel):
    motivo: str = ""
    observaciones: str = ""

# Endpoints Tracto
@app.get("/api/proyectos/tractos")
def get_tractos():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Tracto, Nombre_Tracto FROM Tbl_Proyectos_Tracto ORDER BY Nombre_Tracto")
        rows = cursor.fetchall()
        return [{"id": r[0], "nombre": r[1]} for r in rows]
    finally:
        conn.close()

@app.post("/api/proyectos/tractos")
def add_tracto(payload: TractoPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Tbl_Proyectos_Tracto (Nombre_Tracto) VALUES (?)", (payload.nombre.upper(),))
        conn.commit()
        return {"status": "success"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"El Tracto '{payload.nombre}' ya existe o hay un error de integridad. Detalle: {str(e)}")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno del servidor al insertar tracto: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/proyectos/tractos/{id_tracto}")
def delete_tracto(id_tracto: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_Proyectos_Tracto WHERE ID_Tracto = ?", (id_tracto,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado")
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()

# Endpoints Tipo de Proyecto
@app.get("/api/proyectos/tipos/{id_tracto}")
def get_tipos(id_tracto: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Tipo, Nombre_Tipo FROM Tbl_Tipos_Proyecto WHERE ID_Tracto = ? ORDER BY Nombre_Tipo", (id_tracto,))
        rows = cursor.fetchall()
        return [{"id": r[0], "nombre": r[1]} for r in rows]
    finally:
        conn.close()

@app.post("/api/proyectos/tipos")
def add_tipo(payload: TipoProyectoPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Tbl_Tipos_Proyecto (ID_Tracto, Nombre_Tipo) VALUES (?, ?)", (payload.id_tracto, payload.nombre.upper()))
        conn.commit()
        return {"status": "success"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Error al agregar Tipo. Verifica que no exista ya y que el Tracto sea válido. Detalle: {str(e)}")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/proyectos/tipos/{id_tipo}")
def delete_tipo(id_tipo: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_Tipos_Proyecto WHERE ID_Tipo = ?", (id_tipo,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado")
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()

# Endpoints Version
@app.get("/api/proyectos/versiones/{id_tipo}")
def get_versiones(id_tipo: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Version, Nombre_Version FROM Tbl_Versiones_Ingenieria WHERE ID_Tipo = ? ORDER BY Nombre_Version", (id_tipo,))
        rows = cursor.fetchall()
        return [{"id": r[0], "nombre": r[1]} for r in rows]
    finally:
        conn.close()

@app.post("/api/proyectos/versiones")
def add_version(payload: VersionPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Tbl_Versiones_Ingenieria (ID_Tipo, Nombre_Version) VALUES (?, ?)", (payload.id_tipo, payload.nombre.upper()))
        conn.commit()
        return {"status": "success"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Error en inserción de Versión. Asegúrate de que el Tipo de Proyecto exista. Detalle: {str(e)}")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno del servidor al insertar versión: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/proyectos/versiones/{id_version}")
def delete_version(id_version: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_Versiones_Ingenieria WHERE ID_Version = ?", (id_version,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado")
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()

# Endpoints Clientes
@app.get("/api/proyectos/clientes/{id_version}")
def get_clientes(id_version: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Config_Cliente, Nombre_Cliente FROM Tbl_Clientes_Configuracion WHERE ID_Version = ? ORDER BY Nombre_Cliente", (id_version,))
        rows = cursor.fetchall()
        return [{"id": r[0], "nombre": r[1]} for r in rows]
    finally:
        conn.close()

@app.post("/api/proyectos/clientes")
def add_cliente(payload: ClientePayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Tbl_Clientes_Configuracion (ID_Version, Nombre_Cliente) VALUES (?, ?)", (payload.id_version, payload.nombre.upper()))
        conn.commit()
        return {"status": "success"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Error en inserción de Cliente. Asegúrate de que la Versión exista. Detalle: {str(e)}")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno del servidor al insertar cliente: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/proyectos/clientes/{id_cliente}")
def delete_cliente(id_cliente: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_Clientes_Configuracion WHERE ID_Config_Cliente = ?", (id_cliente,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="No encontrado")
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()
# === FIN JERARQUIA DE PROYECTOS ===

# === MAPA DE INGENIERÍA ===
@app.get("/api/mapa/jerarquia")
def get_mapa_jerarquia():
    """Devuelve el árbol completo: Tracto > Tipo > Versión > Revisiones."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT
                TR.ID_Tracto, TR.Nombre_Tracto,
                TP.ID_Tipo, TP.Nombre_Tipo,
                V.ID_Version, V.Nombre_Version,
                R.ID_Revision, R.Numero_Revision, R.Estado, R.Fecha_Creacion
            FROM Tbl_Proyectos_Tracto TR
            JOIN Tbl_Tipos_Proyecto TP ON TP.ID_Tracto = TR.ID_Tracto
            JOIN Tbl_Versiones_Ingenieria V ON V.ID_Tipo = TP.ID_Tipo
            LEFT JOIN Tbl_BOM_Revisiones R ON R.ID_Version = V.ID_Version
            ORDER BY TR.Nombre_Tracto, TP.Nombre_Tipo, V.Nombre_Version, R.Numero_Revision
        """)
        rows = cursor.fetchall()
        # Construir árbol en Python
        tractos: dict = {}
        for r in rows:
            tid = r.ID_Tracto
            if tid not in tractos:
                tractos[tid] = {"id": tid, "nombre": r.Nombre_Tracto, "tipos": {}}
            tipos = tractos[tid]["tipos"]
            pid = r.ID_Tipo
            if pid not in tipos:
                tipos[pid] = {"id": pid, "nombre": r.Nombre_Tipo, "versiones": {}}
            versiones = tipos[pid]["versiones"]
            vid = r.ID_Version
            if vid not in versiones:
                versiones[vid] = {"id": vid, "nombre": r.Nombre_Version, "revisiones": []}
            if r.ID_Revision:
                versiones[vid]["revisiones"].append({
                    "id_revision": r.ID_Revision,
                    "numero_revision": r.Numero_Revision,
                    "estado": r.Estado,
                    "fecha_creacion": r.Fecha_Creacion.isoformat() if r.Fecha_Creacion else None
                })
        # Serializar a lista
        result = []
        for tracto in tractos.values():
            t = {"id": tracto["id"], "nombre": tracto["nombre"], "tipos": []}
            for tipo in tracto["tipos"].values():
                tp = {"id": tipo["id"], "nombre": tipo["nombre"], "versiones": []}
                for ver in tipo["versiones"].values():
                    tp["versiones"].append({
                        "id": ver["id"],
                        "nombre": ver["nombre"],
                        "revisiones": ver["revisiones"]
                    })
                t["tipos"].append(tp)
            result.append(t)
        return result
    finally:
        conn.close()

# === NUBE DE ARCHIVOS VIN ===
VIN_FILES_BASE = r"C:\BDIV_Archivos\VINs"

@app.get("/api/vins/{id_vin}/archivos")
def get_archivos_vin(id_vin: int):
    """Lista archivos adjuntos de un VIN."""
    folder = os.path.join(VIN_FILES_BASE, str(id_vin))
    if not os.path.exists(folder):
        return []
    archivos = []
    for fname in os.listdir(folder):
        fpath = os.path.join(folder, fname)
        if os.path.isfile(fpath):
            archivos.append({
                "nombre": fname,
                "tamano_kb": round(os.path.getsize(fpath) / 1024, 1),
                "es_pdf": fname.lower().endswith(".pdf")
            })
    return archivos

@app.post("/api/vins/{id_vin}/subir_archivo")
async def subir_archivo_vin(id_vin: int, file: UploadFile = File(...)):
    """Sube y guarda un archivo en la carpeta del VIN en el servidor."""
    folder = os.path.join(VIN_FILES_BASE, str(id_vin))
    os.makedirs(folder, exist_ok=True)
    # Sanitizar nombre de archivo
    safe_name = re.sub(r"[^\w\.\-]", "_", file.filename or "archivo")
    dest = os.path.join(folder, safe_name)
    try:
        content = await file.read()
        with open(dest, "wb") as f:
            f.write(content)
        return {"status": "success", "nombre": safe_name, "tamano_kb": round(len(content) / 1024, 1)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar archivo: {str(e)}")

@app.get("/api/vins/{id_vin}/archivos/{nombre_archivo}")
async def descargar_archivo_vin(id_vin: int, nombre_archivo: str):
    """Descarga/abre un archivo adjunto del VIN."""
    safe_name = re.sub(r"[^\w\.\-]", "_", nombre_archivo)
    fpath = os.path.join(VIN_FILES_BASE, str(id_vin), safe_name)
    if not os.path.exists(fpath):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    def iterfile():
        with open(fpath, "rb") as f:
            yield from f
    media_type = "application/pdf" if safe_name.lower().endswith(".pdf") else "application/octet-stream"
    return StreamingResponse(iterfile(), media_type=media_type,
        headers={"Content-Disposition": f"inline; filename={safe_name}"})


# === HELPER: Registro de Auditoría ===
def registrar_log(cursor, id_revision: int, accion: str, detalle: str, motivo: str = ""):
    """Inserta un registro en Tbl_Log_Cambios_Ingenieria. Llamar dentro de una transacción abierta."""
    try:
        cursor.execute(
            "INSERT INTO Tbl_Log_Cambios_Ingenieria (ID_Revision, Accion, Detalle_Cambio, Motivo) VALUES (?, ?, ?, ?)",
            (id_revision, accion, detalle[:500], motivo[:300] if motivo else "")
        )
    except Exception:
        pass  # No interrumpir operación principal si falla el log

# === MODULO: BOM (Gestor de Listas) ===
@app.get("/api/bom/estaciones/{id_revision}")
def get_estaciones(id_revision: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Estacion, ID_Revision, Nombre_Estacion, Orden FROM Tbl_Estaciones WHERE ID_Revision = ? ORDER BY Orden", (id_revision,))
        rows = cursor.fetchall()
        return [{"id": r.ID_Estacion, "id_revision": r.ID_Revision, "nombre": r.Nombre_Estacion, "orden": r.Orden} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener estaciones: {str(e)}")
    finally:
        conn.close()

@app.post("/api/bom/estaciones")
# Endpoints Revisiones
# --- Endpoints Revisiones (v60.0: agrupados por ID_Version, no por cliente) ---
@app.get("/api/bom/revisiones/version/{id_version}")
def get_revisiones_por_version(id_version: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT ID_Revision, Numero_Revision, Estado, Fecha_Creacion FROM Tbl_BOM_Revisiones WHERE ID_Version = ? ORDER BY Numero_Revision",
            (id_version,)
        )
        rows = cursor.fetchall()
        return [{"id_revision": r.ID_Revision, "numero_revision": r.Numero_Revision, "estado": r.Estado, "fecha_creacion": r.Fecha_Creacion.isoformat() if r.Fecha_Creacion else None} for r in rows]
    finally:
        conn.close()

@app.post("/api/bom/revisiones/version/{id_version}")
def add_revision_version(id_version: int, payload: RevisionPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ISNULL(MAX(Numero_Revision), -1) + 1 FROM Tbl_BOM_Revisiones WHERE ID_Version = ?", (id_version,))
        siguiente_rev = int(cursor.fetchone()[0])
        cursor.execute(
            "INSERT INTO Tbl_BOM_Revisiones (ID_Version, Numero_Revision, Estado) OUTPUT INSERTED.ID_Revision VALUES (?, ?, 'Borrador')",
            (id_version, siguiente_rev)
        )
        id_rev = cursor.fetchone()[0]
        registrar_log(cursor, id_rev, "Creación", f"Nueva Revisión {siguiente_rev} creada para Versión ID {id_version}")
        conn.commit()
        return {"status": "success", "id_revision": id_rev}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error creando revisión: {str(e)}")
    finally:
        conn.close()

# --- Compatibilidad legacy: revisiones por cliente (redirige a versión) ---
@app.get("/api/bom/revisiones/{id_cliente}")
def get_revisiones(id_cliente: int):
    """Legacy endpoint - mantiene compatibilidad con pantallas antiguas."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Buscar revisiones cuya versión corresponde al cliente
        cursor.execute(
            """
            SELECT R.ID_Revision, R.Numero_Revision, R.Estado, R.Fecha_Creacion
            FROM Tbl_BOM_Revisiones R
            JOIN Tbl_Clientes_Configuracion CC ON CC.ID_Version = R.ID_Version
            WHERE CC.ID_Config_Cliente = ?
            ORDER BY R.Numero_Revision
            """,
            (id_cliente,)
        )
        rows = cursor.fetchall()
        return [{"id_revision": r.ID_Revision, "numero_revision": r.Numero_Revision, "estado": r.Estado, "fecha_creacion": r.Fecha_Creacion.isoformat() if r.Fecha_Creacion else None} for r in rows]
    finally:
        conn.close()

@app.post("/api/bom/revisiones/{id_cliente}")
def add_revision(id_cliente: int, payload: RevisionPayload):
    """Legacy endpoint."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Version FROM Tbl_Clientes_Configuracion WHERE ID_Config_Cliente = ?", (id_cliente,))
        ver_row = cursor.fetchone()
        if not ver_row:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        id_version = ver_row[0]
        cursor.execute("SELECT ISNULL(MAX(Numero_Revision), -1) + 1 FROM Tbl_BOM_Revisiones WHERE ID_Version = ?", (id_version,))
        siguiente_rev = int(cursor.fetchone()[0])
        cursor.execute(
            "INSERT INTO Tbl_BOM_Revisiones (ID_Version, Numero_Revision, Estado) OUTPUT INSERTED.ID_Revision VALUES (?, ?, 'Borrador')",
            (id_version, siguiente_rev)
        )
        id_rev = cursor.fetchone()[0]
        registrar_log(cursor, id_rev, "Creación", f"Nueva Revisión {siguiente_rev} (vía cliente {id_cliente})")
        conn.commit()
        return {"status": "success", "id_revision": id_rev}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error creando revisión: {str(e)}")
    finally:
        conn.close()

@app.put("/api/bom/revisiones/{id_revision}/aprobar")
def aprobar_revision(id_revision: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE Tbl_BOM_Revisiones SET Estado = 'Aprobada' WHERE ID_Revision = ?", (id_revision,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Revisión no encontrada")
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()

# Endpoints VINs (Unidades Físicas)
@app.get("/api/bom/revisiones/{id_revision}/vins")
def get_vins(id_revision: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Unidad, VIN FROM Tbl_Unidades_Fisicas WHERE ID_Revision = ?", (id_revision,))
        rows = cursor.fetchall()
        return [{"id_unidad": r.ID_Unidad, "vin": r.VIN} for r in rows]
    finally:
        conn.close()

@app.get("/api/bom/buscar_pieza_jerarquia/{codigo_pieza}")
def buscar_pieza_jerarquia(codigo_pieza: str, exclude_rev: Optional[int] = None):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        query = """
            SELECT 
                R.ID_Revision, 
                TR.Nombre_Tracto, 
                TP.Nombre_Tipo, 
                V.Nombre_Version, 
                CC.Nombre_Cliente, 
                R.Numero_Revision, 
                R.Estado, 
                E.Cantidad
            FROM Tbl_BOM_Estructura E
            JOIN Tbl_Ensambles EN ON E.ID_Ensamble = EN.ID_Ensamble
            JOIN Tbl_Estaciones ES ON EN.ID_Estacion = ES.ID_Estacion
            JOIN Tbl_BOM_Revisiones R ON ES.ID_Revision = R.ID_Revision
            JOIN Tbl_Versiones_Ingenieria V ON R.ID_Version = V.ID_Version
            LEFT JOIN Tbl_Clientes_Configuracion CC ON CC.ID_Version = V.ID_Version
            JOIN Tbl_Tipos_Proyecto TP ON V.ID_Tipo = TP.ID_Tipo
            JOIN Tbl_Proyectos_Tracto TR ON TP.ID_Tracto = TR.ID_Tracto
            WHERE E.Codigo_Pieza = ?
        """
        params = [codigo_pieza]
        if exclude_rev:
            query += " AND R.ID_Revision != ?"
            params.append(exclude_rev)
            
        cursor.execute(query, params)
        rows = cursor.fetchall()
        return [
            {
                "id_revision": r.ID_Revision,
                "tracto": r.Nombre_Tracto,
                "tipo": r.Nombre_Tipo,
                "version": r.Nombre_Version,
                "cliente": r.Nombre_Cliente,
                "numero_revision": r.Numero_Revision,
                "estado": r.Estado,
                "cantidad": r.Cantidad
            } for r in rows
        ]
    finally:
        conn.close()

@app.get("/api/bom/exportar/{id_revision}")
def exportar_bom(id_revision: int):
    """Exportación Pro: BOM + Historial de Auditoría en 2 pestañas."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Pestaña 1: BOM
        cursor.execute(
            """
            SELECT ES.Nombre_Estacion, EN.Nombre_Ensamble, E.Codigo_Pieza,
                   M.Descripcion as Descripcion_Oficial, E.Cantidad
            FROM Tbl_BOM_Estructura E
            JOIN Tbl_Ensambles EN ON E.ID_Ensamble = EN.ID_Ensamble
            JOIN Tbl_Estaciones ES ON EN.ID_Estacion = ES.ID_Estacion
            LEFT JOIN Tbl_Maestro_Piezas M ON E.Codigo_Pieza = M.Codigo_Pieza
            WHERE ES.ID_Revision = ?
            ORDER BY ES.Orden, EN.Nombre_Ensamble
            """, (id_revision,)
        )
        bom_rows = cursor.fetchall()

        # Pestaña 2: Historial de Auditoría
        try:
            cursor.execute(
                "SELECT Usuario, Fecha_Hora, Accion, Detalle_Cambio, Motivo FROM Tbl_Log_Cambios_Ingenieria WHERE ID_Revision = ? ORDER BY Fecha_Hora DESC",
                (id_revision,)
            )
            log_rows = cursor.fetchall()
        except Exception:
            log_rows = []

        output = io.BytesIO()
        workbook = openpyxl.Workbook()

        # --- Pestaña 1: BOM (formato cliente: cabecera fila 5, datos desde fila 6) ---
        sheet_bom = workbook.active
        sheet_bom.title = "BOM"
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_align = Alignment(horizontal="center")

        # Cabecera en fila 5, columnas B/C/D/G (índices 2,3,4,7)
        COL_ESTACION  = 2  # B
        COL_ENSAMBLE  = 3  # C
        COL_CODIGO    = 4  # D
        COL_CANTIDAD  = 7  # G
        HDR_ROW = 5
        DATA_ROW_START = 6

        headers = {COL_ESTACION: "Estación", COL_ENSAMBLE: "Ensamble",
                   COL_CODIGO: "Código Pieza", COL_CANTIDAD: "Cantidad"}
        for col_idx, text in headers.items():
            cell = sheet_bom.cell(row=HDR_ROW, column=col_idx, value=text)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

        for row_offset, r in enumerate(bom_rows):
            row_num = DATA_ROW_START + row_offset
            sheet_bom.cell(row=row_num, column=COL_ESTACION,  value=r.Nombre_Estacion)
            sheet_bom.cell(row=row_num, column=COL_ENSAMBLE,  value=r.Nombre_Ensamble)
            sheet_bom.cell(row=row_num, column=COL_CODIGO,    value=r.Codigo_Pieza)
            sheet_bom.cell(row=row_num, column=COL_CANTIDAD,  value=r.Cantidad)

        # Ancho de columnas usadas
        for col_idx, ancho in [(COL_ESTACION, 22), (COL_ENSAMBLE, 28),
                               (COL_CODIGO, 18), (COL_CANTIDAD, 10)]:
            col_letter = sheet_bom.cell(row=1, column=col_idx).column_letter
            sheet_bom.column_dimensions[col_letter].width = ancho

        # --- Pestaña 2: Historial ---
        sheet_log = workbook.create_sheet(title="Historial de Cambios")
        log_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        log_headers = ["Usuario", "Fecha / Hora", "Acción", "Detalle", "Motivo"]
        sheet_log.append(log_headers)
        for cell in sheet_log[1]:
            cell.font = hdr_font; cell.fill = log_fill; cell.alignment = hdr_align
        for r in log_rows:
            sheet_log.append([
                r.Usuario,
                r.Fecha_Hora.strftime("%Y-%m-%d %H:%M:%S") if r.Fecha_Hora else '',
                r.Accion, r.Detalle_Cambio, r.Motivo or ''
            ])

        workbook.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=BOM_Rev{id_revision}_v60.xlsx"}
        )
    finally:
        conn.close()

@app.get("/api/bom/log/{id_revision}")
def get_log_auditoria(id_revision: int):
    """ADN de Ingeniería: historial completo de cambios de una revisión."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT ID_Log, Usuario, Fecha_Hora, Accion, Detalle_Cambio, Motivo FROM Tbl_Log_Cambios_Ingenieria WHERE ID_Revision = ? ORDER BY Fecha_Hora DESC",
            (id_revision,)
        )
        rows = cursor.fetchall()
        return [{
            "id_log": r.ID_Log,
            "usuario": r.Usuario,
            "fecha_hora": r.Fecha_Hora.isoformat() if r.Fecha_Hora else None,
            "accion": r.Accion,
            "detalle": r.Detalle_Cambio,
            "motivo": r.Motivo or ""
        } for r in rows]
    except Exception:
        return []  # Si la tabla aún no existe, retorna lista vacía
    finally:
        conn.close()

@app.put("/api/proyectos/clientes/{id_cliente}/asignar_revision")
def asignar_revision_cliente(id_cliente: int, payload: AsignarRevisionPayload):
    """Vincula un cliente a una revisión maestra específica."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE Tbl_Clientes_Configuracion SET ID_Revision_Asignada = ? WHERE ID_Config_Cliente = ?",
            (payload.id_revision_asignada, id_cliente)
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.post("/api/bom/revisiones/{id_revision}/vins")
def add_vin(id_revision: int, payload: VINPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Tbl_Unidades_Fisicas (ID_Revision, VIN) OUTPUT INSERTED.ID_Unidad VALUES (?, ?)", (id_revision, payload.vin.upper()))
        id_gen = cursor.fetchone()[0]
        conn.commit()
        return {"status": "success", "id_unidad": id_gen}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error al agregar VIN: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/bom/vins/{id_unidad}")
def delete_vin(id_unidad: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_Unidades_Fisicas WHERE ID_Unidad = ?", (id_unidad,))
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()

# --- FUNCIONALIDADES AVANZADAS (FASE 8) ---

@app.get("/api/vins/buscar")
def buscar_vin(q: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        query = """
            SELECT u.ID_Unidad, u.VIN, u.Notas, r.ID_Revision, r.Numero_Revision,
                   v.ID_Version, c.ID_Config_Cliente, c.Nombre_Cliente,
                   v.Nombre_Version, t.Nombre_Tipo, tr.Nombre_Tracto
            FROM Tbl_Unidades_Fisicas u
            JOIN Tbl_BOM_Revisiones r ON u.ID_Revision = r.ID_Revision
            JOIN Tbl_Versiones_Ingenieria v ON r.ID_Version = v.ID_Version
            LEFT JOIN Tbl_Clientes_Configuracion c ON c.ID_Version = v.ID_Version
            JOIN Tbl_Tipos_Proyecto t ON v.ID_Tipo = t.ID_Tipo
            JOIN Tbl_Proyectos_Tracto tr ON t.ID_Tracto = tr.ID_Tracto
            WHERE u.VIN LIKE ?
        """
        cursor.execute(query, (f"%{q}%",))
        rows = cursor.fetchall()
        return [
            {
                "id_unidad": r.ID_Unidad,
                "vin": r.VIN,
                "notas": r.Notas,
                "id_revision": r.ID_Revision,
                "numero_revision": r.Numero_Revision,
                "id_cliente": r.ID_Config_Cliente,
                "cliente": r.Nombre_Cliente,
                "version": r.Nombre_Version,
                "tipo": r.Nombre_Tipo,
                "tracto": r.Nombre_Tracto
            } for r in rows
        ]
    finally:
        conn.close()

@app.put("/api/vins/{id_unidad}/notas")
def update_vin_notas(id_unidad: int, payload: VINPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE Tbl_Unidades_Fisicas SET Notas = ? WHERE ID_Unidad = ?", (payload.notas, id_unidad))
        conn.commit()
        return {"status": "success"}
    finally:
        conn.close()

@app.post("/api/bom/clonar")
def clonar_bom(payload: ClonarPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # 1. Obtener estaciones del origen
        cursor.execute("SELECT ID_Estacion, Nombre_Estacion, Orden FROM Tbl_Estaciones WHERE ID_Revision = ?", (payload.id_revision_origen,))
        estaciones = cursor.fetchall()
        
        for est_orig in estaciones:
            id_est_orig = est_orig.ID_Estacion
            # Insertar nueva estación
            cursor.execute(
                "INSERT INTO Tbl_Estaciones (ID_Revision, Nombre_Estacion, Orden) OUTPUT INSERTED.ID_Estacion VALUES (?, ?, ?)",
                (payload.id_revision_destino, est_orig.Nombre_Estacion, est_orig.Orden)
            )
            id_est_dest = cursor.fetchone()[0]
            
            # 2. Obtener ensambles de la estación origen
            cursor.execute("SELECT ID_Ensamble, Nombre_Ensamble FROM Tbl_Ensambles WHERE ID_Estacion = ?", (id_est_orig,))
            ensambles = cursor.fetchall()
            
            for ens_orig in ensambles:
                id_ens_orig = ens_orig.ID_Ensamble
                # Insertar nuevo ensamble
                cursor.execute(
                    "INSERT INTO Tbl_Ensambles (ID_Estacion, Nombre_Ensamble) OUTPUT INSERTED.ID_Ensamble VALUES (?, ?)",
                    (id_est_dest, ens_orig.Nombre_Ensamble)
                )
                id_ens_dest = cursor.fetchone()[0]
                
                # 3. Obtener piezas del ensamble origen
                cursor.execute("SELECT Codigo_Pieza, Cantidad, Observaciones FROM Tbl_BOM_Estructura WHERE ID_Ensamble = ?", (id_ens_orig,))
                piezas = cursor.fetchall()
                
                for p in piezas:
                    cursor.execute(
                        "INSERT INTO Tbl_BOM_Estructura (ID_Ensamble, Codigo_Pieza, Cantidad, Observaciones) VALUES (?, ?, ?, ?)",
                        (id_ens_dest, p.Codigo_Pieza, p.Cantidad, p.Observaciones)
                    )
        
        conn.commit()
        return {"status": "success", "detalle": f"Clonadas {len(estaciones)} estaciones."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error al clonar: {str(e)}")
    finally:
        conn.close()

@app.post("/api/bom/propagar")
def propagar_cambios(payload: PropagarPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if not payload.id_revisiones:
            return {"status": "ignored", "mensaje": "No se seleccionaron revisiones"}
            
        # Generar placeholders para la lista de IDs
        placeholders = ",".join(["?"] * len(payload.id_revisiones))
        query = f"""
            UPDATE be
            SET be.Cantidad = ?
            FROM Tbl_BOM_Estructura be
            JOIN Tbl_Ensambles en ON be.ID_Ensamble = en.ID_Ensamble
            JOIN Tbl_Estaciones es ON en.ID_Estacion = es.ID_Estacion
            WHERE be.Codigo_Pieza = ? AND es.ID_Revision IN ({placeholders})
        """
        params = [payload.nueva_cantidad, payload.codigo_pieza] + payload.id_revisiones
        cursor.execute(query, params)
        affected = cursor.rowcount
        conn.commit()
        return {"status": "success", "afectados": affected}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en propagación: {str(e)}")
    finally:
        conn.close()

@app.get("/api/bom/estaciones/{id_revision}")
def get_estaciones(id_revision: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Estacion, ID_Revision, Nombre_Estacion, Orden FROM Tbl_Estaciones WHERE ID_Revision = ? ORDER BY Orden", (id_revision,))
        rows = cursor.fetchall()
        return [{"id": r.ID_Estacion, "id_revision": r.ID_Revision, "nombre": r.Nombre_Estacion, "orden": r.Orden} for r in rows]
    finally:
        conn.close()

@app.post("/api/bom/estaciones")
def add_estacion(payload: EstacionPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Tbl_Estaciones (ID_Revision, Nombre_Estacion, Orden) VALUES (?, ?, ?)", (payload.id_revision, payload.nombre.upper(), 0))
        conn.commit()
        return {"status": "success"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Error al agregar Estación. Asegúrate de que no exista duplicada. Detalle: {str(e)}")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL interno en Estación: {str(e)}")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/bom/estaciones/{id_estacion}")
def delete_estacion(id_estacion: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Borrar piezas de sus ensambles
        cursor.execute("SELECT ID_Ensamble FROM Tbl_Ensambles WHERE ID_Estacion = ?", (id_estacion,))
        ensambles = cursor.fetchall()
        for ens in ensambles:
            cursor.execute("DELETE FROM Tbl_BOM_Estructura WHERE ID_Ensamble = ?", (ens.ID_Ensamble,))
        
        # Borrar los ensambles
        cursor.execute("DELETE FROM Tbl_Ensambles WHERE ID_Estacion = ?", (id_estacion,))
        
        # Borrar la estacion
        cursor.execute("DELETE FROM Tbl_Estaciones WHERE ID_Estacion = ?", (id_estacion,))
        conn.commit()
        return {"status": "success"}
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL al eliminar estación: {str(e)}")
    finally:
        conn.close()

@app.get("/api/bom/ensambles/{id_estacion}")
def get_ensambles(id_estacion: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Ensamble, ID_Estacion, Codigo_Ensamble, Nombre_Ensamble FROM Tbl_Ensambles WHERE ID_Estacion = ? ORDER BY Nombre_Ensamble", (id_estacion,))
        rows = cursor.fetchall()
        return [{"id": r.ID_Ensamble, "id_estacion": r.ID_Estacion, "codigo_ensamble": r.Codigo_Ensamble, "nombre": r.Nombre_Ensamble} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener ensambles: {str(e)}")
    finally:
        conn.close()

@app.post("/api/bom/ensambles")
def add_ensamble(payload: EnsamblePayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Se asume un Codigo_Ensamble generico por ahora
        cursor.execute("INSERT INTO Tbl_Ensambles (ID_Estacion, Codigo_Ensamble, Nombre_Ensamble) VALUES (?, ?, ?)", (payload.id_estacion, "N/A", payload.nombre.upper()))
        conn.commit()
        return {"status": "success"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Error de integridad en Ensamble. Detalle: {str(e)}")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL en Ensamble: {str(e)}")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    finally:
        conn.close()

@app.delete("/api/bom/ensambles/{id_ensamble}")
def delete_ensamble(id_ensamble: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Borrar piezas
        cursor.execute("DELETE FROM Tbl_BOM_Estructura WHERE ID_Ensamble = ?", (id_ensamble,))
        
        # Borrar ensamble
        cursor.execute("DELETE FROM Tbl_Ensambles WHERE ID_Ensamble = ?", (id_ensamble,))
        conn.commit()
        return {"status": "success"}
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL al eliminar ensamble: {str(e)}")
    finally:
        conn.close()

@app.get("/api/bom/estructura/{id_ensamble}")
def get_bom_estructura(id_ensamble: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT e.ID_BOM, e.Codigo_Pieza, m.Descripcion, e.Cantidad, e.Observaciones_Proceso
            FROM Tbl_BOM_Estructura e
            LEFT JOIN Tbl_Maestro_Piezas m ON e.Codigo_Pieza = m.Codigo_Pieza
            WHERE e.ID_Ensamble = ?
        """, (id_ensamble,))
        rows = cursor.fetchall()
        return [{
            "id": r.ID_BOM,
            "codigo": r.Codigo_Pieza,
            "descripcion": getattr(r, 'Descripcion', 'Descripción no encontrada') if getattr(r, 'Descripcion', None) else "Descripción no encontrada",
            "cantidad": r.Cantidad,
            "observaciones": r.Observaciones_Proceso or ""
        } for r in rows]
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=f"Error SQL al obtener estructura: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno al obtener estructura: {str(e)}")
    finally:
        conn.close()

@app.post("/api/bom/estructura")
def add_bom_estructura(payload: BOMPayload):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO Tbl_BOM_Estructura (ID_Ensamble, Codigo_Pieza, Cantidad, Observaciones_Proceso)
            VALUES (?, ?, ?, ?)
        """, (payload.id_ensamble, payload.codigo_pieza, payload.cantidad, payload.observaciones))
        conn.commit()
        return {"status": "success"}
    except pyodbc.IntegrityError as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Error de integridad en BOM. Verifica Código existete: {str(e)}")
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL en BOM_Estructura: {str(e)}")
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.delete("/api/bom/estructura/{id_bom}")
def delete_bom_estructura(id_bom: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Tbl_BOM_Estructura WHERE ID_BOM = ?", (id_bom,))
        conn.commit()
        return {"status": "success"}
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL al eliminar BOM_Estructura: {str(e)}")
    finally:
        conn.close()

class BOMPiezaUpdate(BaseModel):
    cantidad: float

@app.put("/api/bom/piezas/{id_bom}")
def update_bom_pieza(id_bom: int, payload: BOMPiezaUpdate):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Validación de la cantidad
        if payload.cantidad <= 0:
            raise HTTPException(status_code=400, detail="La cantidad debe ser mayor a 0.")
            
        cursor.execute("UPDATE Tbl_BOM_Estructura SET Cantidad = ? WHERE ID_BOM = ?", (payload.cantidad, id_bom))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Pieza de BOM no encontrada.")
        conn.commit()
        return {"status": "success"}
    except pyodbc.Error as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL al actualizar la pieza: {str(e)}")
    finally:
        conn.close()

@app.get("/api/bom/arbol/{id_revision}")
def get_bom_arbol(id_revision: int):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT ID_Estacion, ID_Revision, Nombre_Estacion, Orden FROM Tbl_Estaciones WHERE ID_Revision = ? ORDER BY Orden", (id_revision,))
        estaciones = cursor.fetchall()
        
        arbol = []
        for est in estaciones:
            id_estacion = est.ID_Estacion
            est_dict = {"id": id_estacion, "nombre": est.Nombre_Estacion, "ensambles": []}
            
            cursor.execute("SELECT ID_Ensamble, ID_Estacion, Codigo_Ensamble, Nombre_Ensamble FROM Tbl_Ensambles WHERE ID_Estacion = ? ORDER BY Nombre_Ensamble", (id_estacion,))
            ensambles = cursor.fetchall()
            
            for ens in ensambles:
                id_ensamble = ens.ID_Ensamble
                ens_dict = {"id": id_ensamble, "nombre": ens.Nombre_Ensamble, "piezas": []}
                
                cursor.execute("""
                    SELECT e.ID_BOM, e.Codigo_Pieza, m.Descripcion, e.Cantidad, e.Observaciones_Proceso,
                           m.Simetria, m.Proceso_Primario, m.Proceso_1, m.Proceso_2, m.Proceso_3, m.Link_Drive
                    FROM Tbl_BOM_Estructura e
                    LEFT JOIN Tbl_Maestro_Piezas m ON e.Codigo_Pieza = m.Codigo_Pieza
                    WHERE e.ID_Ensamble = ?
                """, (id_ensamble,))
                piezas = cursor.fetchall()
                
                for p_row in piezas:
                    desc = getattr(p_row, 'Descripcion', None)
                    ens_dict["piezas"].append({
                        "id": p_row.ID_BOM,
                        "codigo": p_row.Codigo_Pieza,
                        "descripcion": desc if desc else "N/A",
                        "cantidad": p_row.Cantidad,
                        "observaciones": p_row.Observaciones_Proceso or "",
                        "simetria": getattr(p_row, 'Simetria', '') or "",
                        "proceso_primario": getattr(p_row, 'Proceso_Primario', '') or "",
                        "proceso_1": getattr(p_row, 'Proceso_1', '') or "",
                        "proceso_2": getattr(p_row, 'Proceso_2', '') or "",
                        "proceso_3": getattr(p_row, 'Proceso_3', '') or "",
                        "link_drive": getattr(p_row, 'Link_Drive', '') or ""
                    })
                
                est_dict["ensambles"].append(ens_dict)
                
            arbol.append(est_dict)
            
        return arbol
    except pyodbc.Error as e:
        raise HTTPException(status_code=500, detail=f"Error SQL en el Árbol BOM: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno en Árbol BOM: {str(e)}")
    finally:
        conn.close()
@app.post("/api/bom/importar/{id_revision}")
async def importar_bom(id_revision: int, file: UploadFile = File(...)):
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx, .xls)")
    
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content), header=None)
    except Exception as e:
        if isinstance(e, HTTPException): raise e
        raise HTTPException(status_code=400, detail=f"Error al analizar el Excel: {str(e)}")
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    total_leidos = 0
    insertados = 0
    errores_mapeo = []
    
    try:
        # Pre-cargar catálogo maestro para validación rápida
        cursor.execute("SELECT Codigo_Pieza FROM Tbl_Maestro_Piezas")
        codigos_buscar = {str(row[0]).strip().upper() for row in cursor.fetchall() if row[0]}
        
        acumulados = {}
        
        # 1.- Lógica de Secuencia Inicial (Ensambles)
        cursor.execute("SELECT MAX(Codigo_Ensamble) FROM Tbl_Ensambles WHERE Codigo_Ensamble LIKE 'E-%'")
        max_code_row = cursor.fetchone()
        secuencia_ensamble = 0
        if max_code_row and max_code_row[0]:
            try:
                # Extraer número tras el guión (ej: E-0005 -> 5)
                secuencia_ensamble = int(max_code_row[0].split('-')[1])
            except (ValueError, IndexError):
                pass

        for index, row in df.iterrows():
            estacion_val = row[1]
            ensamble_val = row[2]
            codigo_val = row[3]
            cantidad_val = row[6] if len(row) > 6 else 1

            if pd.isna(codigo_val) or str(codigo_val).strip() == '' or str(codigo_val).strip().lower() == 'código de pieza':
                continue

            if pd.isna(estacion_val) or pd.isna(ensamble_val):
                continue
                
            total_leidos += 1
                
            estacion_nombre = "" if pd.isna(row[1]) else str(row[1]).strip()
            ensamble_nombre = "" if pd.isna(row[2]) else str(row[2]).strip()
            codigo = str(codigo_val).strip().upper()
            
            # 2. Validación de Existencia
            if codigo not in codigos_buscar:
                if codigo not in errores_mapeo:
                    errores_mapeo.append(codigo)
                continue
            
            try:
                if pd.isna(cantidad_val) or cantidad_val is None:
                    cantidad = 1
                else:
                    cantidad = int(float(cantidad_val))
            except (ValueError, TypeError):
                cantidad = 1
                
            # 3. Acumulación
            llave = (estacion_nombre, ensamble_nombre, codigo)
            acumulados[llave] = acumulados.get(llave, 0) + cantidad
            
        # Inserción final con Caché para velocidad
        cache_estaciones = {}
        cache_ensambles = {}
        
        for (estacion_nombre, ensamble_nombre, codigo), cantidad in acumulados.items():
            # 1. Buscar o Crear ESTACION
            if estacion_nombre not in cache_estaciones:
                cursor.execute("SELECT ID_Estacion FROM Tbl_Estaciones WHERE ID_Revision = ? AND Nombre_Estacion = ?", (id_revision, estacion_nombre))
                est_row = cursor.fetchone()
                if est_row:
                    id_estacion = est_row[0]
                else:
                    cursor.execute("SELECT ISNULL(MAX(Orden), 0) + 1 FROM Tbl_Estaciones WHERE ID_Revision = ?", (id_revision,))
                    nuevo_orden = cursor.fetchone()[0]
                    cursor.execute(
                        "INSERT INTO Tbl_Estaciones (ID_Revision, Nombre_Estacion, Orden) OUTPUT INSERTED.ID_Estacion VALUES (?, ?, ?)", 
                        (id_revision, estacion_nombre, nuevo_orden)
                    )
                    id_estacion = int(cursor.fetchone()[0])
                cache_estaciones[estacion_nombre] = id_estacion
            else:
                id_estacion = cache_estaciones[estacion_nombre]
                
            # 2. Buscar o Crear ENSAMBLE
            ensamble_key = (id_estacion, ensamble_nombre)
            if ensamble_key not in cache_ensambles:
                cursor.execute("SELECT ID_Ensamble FROM Tbl_Ensambles WHERE ID_Estacion = ? AND Nombre_Ensamble = ?", (id_estacion, ensamble_nombre))
                ens_row = cursor.fetchone()
                if ens_row:
                    id_ensamble = ens_row[0]
                else:
                    secuencia_ensamble += 1
                    codigo_ensamble_generado = f"E-{secuencia_ensamble:04d}"
                    cursor.execute(
                        "INSERT INTO Tbl_Ensambles (ID_Estacion, Codigo_Ensamble, Nombre_Ensamble) OUTPUT INSERTED.ID_Ensamble VALUES (?, ?, ?)", 
                        (id_estacion, codigo_ensamble_generado, ensamble_nombre)
                    )
                    id_ensamble = int(cursor.fetchone()[0])
                cache_ensambles[ensamble_key] = id_ensamble
            else:
                id_ensamble = cache_ensambles[ensamble_key]
                
            # 3. Insertar PIEZA (BOM)
            cursor.execute("INSERT INTO Tbl_BOM_Estructura (ID_Ensamble, Codigo_Pieza, Cantidad, Observaciones_Proceso) VALUES (?, ?, ?, ?)", (id_ensamble, codigo, cantidad, ""))
            insertados += 1
            
        conn.commit()
        return {"status": "success", "total_leidos": total_leidos, "insertados": insertados, "errores": errores_mapeo}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error SQL durante importación, transacción revertida: {str(e)}")
    finally:
        conn.close()

# === FIN BOM ===

# === FIN MATERIALES APROBADOS ===
# --- ENDPOINTS CONFIGURACIÓN ---
@app.get("/api/config/regla_espejo")
async def get_mirror_config():
    return {"activa": REGLA_ESPEJO_ACTIVA}

@app.post("/api/config/regla_espejo")
async def set_mirror_config(config: MirrorConfig):
    global REGLA_ESPEJO_ACTIVA
    REGLA_ESPEJO_ACTIVA = config.activa
    print(f"--- REGLA ESPEJO ACTUALIZADA: {REGLA_ESPEJO_ACTIVA} ---")
    return {"activa": REGLA_ESPEJO_ACTIVA}

# Middleware CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db_connection():
    try:
        conn = pyodbc.connect(CONNECTION_STRING)
        return conn
    except Exception as e:
        print(f"Error de conexión SQL: {e}")
        raise HTTPException(status_code=500, detail=f"Database Connection Error: {str(e)}")

def init_auth_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Tbl_Usuarios' AND xtype='U')
            CREATE TABLE Tbl_Usuarios (
                ID INT PRIMARY KEY IDENTITY(1,1),
                Username NVARCHAR(50) UNIQUE NOT NULL,
                Password NVARCHAR(50) NOT NULL,
                Role NVARCHAR(20) DEFAULT 'User'
            )
        """)
        conn.commit()
        
        users_to_seed = [
            ("jaes_admin", "Industrial.2026", "Admin"),
            ("ing_01", "Ing.2026", "User"),
            ("ing_02", "Ing.2026", "User"),
            ("ing_03", "Ing.2026", "User"),
            ("ing_04", "Ing.2026", "User"),
            ("ing_05", "Ing.2026", "User"),
            ("ing_06", "Ing.2026", "User"),
        ]
        
        for user, password, role in users_to_seed:
            cursor.execute("SELECT ID FROM Tbl_Usuarios WHERE Username = ?", (user,))
            if cursor.fetchone():
                cursor.execute("UPDATE Tbl_Usuarios SET Password = ?, Role = ? WHERE Username = ?", (password, role, user))
            else:
                cursor.execute("INSERT INTO Tbl_Usuarios (Username, Password, Role) VALUES (?, ?, ?)", (user, password, role))
        
        conn.commit()
    finally:
        conn.close()

@app.post("/api/login")
def login(request: LoginRequest):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT Role FROM Tbl_Usuarios WHERE Username = ? AND Password = ?", (request.username, request.password))
        user = cursor.fetchone()
        
        conn.close()
        
        if user:
            return {"success": True, "role": user[0]}
        else:
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="Credenciales incorrectas")
            
    except Exception as e:
        print(f"Error en login: {e}")
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="Error interno del servidor")

def iniciar_auditoria():
    """Crea la tabla de auditoría si no existe. No detiene el arranque si falla."""
    print("--- INICIANDO SISTEMA DE AUDITORIA ---")
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Auditoria_Cambios')
            BEGIN
                CREATE TABLE Tbl_Auditoria_Cambios (
                    ID_Log INT IDENTITY(1,1) PRIMARY KEY,
                    Codigo_Pieza VARCHAR(50),
                    Accion VARCHAR(50),
                    Valor_Anterior NVARCHAR(MAX),
                    Valor_Nuevo NVARCHAR(MAX),
                    Usuario VARCHAR(100),
                    Fecha_Hora DATETIME DEFAULT GETDATE()
                );
            END
        """)
        conn.commit()
    # TABLA DE MATERIALES APROBADOS
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Materiales_Aprobados')
            BEGIN
                CREATE TABLE Tbl_Materiales_Aprobados (
                    ID INT IDENTITY(1,1) PRIMARY KEY,
                    Material VARCHAR(200) UNIQUE
                );
            END
        """)
        
        # POBLAR TABLA SI ESTÁ VACÍA
        cursor.execute("SELECT COUNT(*) FROM Tbl_Materiales_Aprobados")
        if cursor.fetchone()[0] == 0:
            materiales_iniciales = [
                "ACERO ASTM A36 1/8\"", "ACERO ASTM A36 3/16\"", "ACERO ASTM A36 C.10", "ACERO ASTM A36 C.14", "ACERO ASTM A36 C.16",
                "ACERO ASTM A572 G50 1/2\"", "ACERO ASTM A572 G50 1/4\"", "ACERO ASTM A572 G50 3/4\"", "ACERO ASTM A572 G50 3/8\"", "ACERO ASTM A572 G50 5/16\"",
                "ACERO INOXIDABLE 304 CAL.16", "ACERO INOXIDABLE C.11", "ALUMINIO 3003 C.11", "ALUMINIO 3003 C.14", "ALUMINIO 5052 1/4\"", "ALUMINIO NEGRO 3003 C.19",
                "ALUMINIO MACIZO 6026 Ø 1 1/2\"", "ALUMINIO MACIZO 6026 Ø 1/2\"", "ALUMINIO MACIZO 6026 Ø 2 1/2\"", "ALUMINIO MACIZO 6026 Ø 2\"", "ALUMINIO MACIZO 6026 Ø 3 1/2\"", "ALUMINIO MACIZO 6026 Ø 3\"", "ALUMINIO MACIZO 6026 Ø 4\"", "ALUMINIO MACIZO 6026 Ø 7/8\"",
                "ANGULO ASTM A36 1 1/2\" x 1 1/2\" x 3/16\"", "ANGULO ASTM A36 1\" x 1\" x 3/16\"", "ANGULO ASTM A36 2\" x 2\" x 3/16\"",
                "BARRA HUECA AISI 1018 Ø 33mm x 14mm", "BARRA HUECA AISI 1018 Ø 40mm x 25mm", "BARRA HUECA AISI 1018 Ø 40mm x 28mm", "BARRA HUECA AISI 1018 Ø 50mm x 35mm", "BARRA HUECA AISI 1018 Ø 76mm x 38mm",
                "BARRA CROMADA AISI 1045 Ø 28mm", "BARRA CROMADA AISI 1045 Ø 45mm", "TUBO HONEADO AISI 1018 Ø 60mm x 50mm", "TUBO HONEADO AISI 1018 Ø 73mm x 63mm",
                "BARRA HUECA CROMADA AISI 1018 Ø 38.1mm X 25.4mm",
                "BARRA HUECA DE ALUMINIO B241 6026 Ø 101.6mm X 50.5mm", "BARRA HUECA DE ALUMINIO B241 6026 Ø 63.5mm X 29.7mm", "BARRA HUECA DE ALUMINIO B241 6026 Ø 76.2mm X 29.7mm", "BARRA HUECA DE ALUMINIO B241 6026 Ø 88.9mm X 24.7mm",
                "CAJA DE TENSADO DE LONA", "CANAL C A36 4\"", "COMERCIAL BISAGRA DE LIBRO", "COMERCIAL BISAGRA DE PIANO", "MATRACA DE LONA", "PERFIL ALUMINIO PELDAÑO 688 6061 T6", "PERNO REY  COMERCIAL", "SEGURO DE FUNDICION -", "SEGURO DE RESORTE CORTO", "SEGURO DE RESORTE LARGO",
                "HSS ASTM A500 °B 2 1/2\" x 2 1/2\" x 1/4\"", "HSS ASTM A500 °B 2 1/2\" x 2 1/2\" x 3/16\"", "HSS ASTM A500 °B 2\" x 2\" x 1/4\"", "HSS ASTM A500 °B 2\" x 2\" x 3/16\"", "HSS ASTM A500 °B 3 1/2\" X 3 1/2\" X 3/16\"", "HSS ASTM A500 °B 3\" x 2\" x 1/4\"", "HSS ASTM A500 °B 3\" x 2\" x 3/16\"", "HSS ASTM A500 °B 3\" x 3\" x 1/4\"", "HSS ASTM A500 °B 3\" x 3\" x 3/16\"", "HSS ASTM A500 °B 4 1/2\" x 3 1/2\" x 3/16\"", "HSS ASTM A500 °B 4\" x 2\" x 1/4\"", "HSS ASTM A500 °B 4\" x 2\" x 3/16\"", "HSS ASTM A500 °B 4\" x 3\" x 1/4\"", "HSS ASTM A500 °B 4\" x 3\" x 3/16\"", "HSS ASTM A500 °B 4\" x 3\" x 3/8\"", "HSS ASTM A500 °B 4\" x 4\" x 3/8\"", "HSS ASTM A500 °B 6\" x 2\" x 1/4\"", "HSS ASTM A500 °B 6\" x 2\" x 3/16\"", "HSS ASTM A500 °B 6\" x 3\" x 1/4\"", "HSS ASTM A500 °B 6\" x 3\" x 3/16\"", "HSS ASTM A500 °B 6\" x 4\" x 1/4\"", "HSS ASTM A500 °B 6\" x 4\" x 3/8\"", "HSS ASTM A500 °B 6\" X 6\" X 1/4\"", "HSS ASTM A500 °B 6\" x 6\" x 3/8\"",
                "PLACA HARDOX 1/4\"", "PLACA STRENX 110 XF 3/16\"", "PLACA STRENX 110XF 1/2\"",
                "PTR ASTM A36 1 1/2\" x 1 1/2 \" x 3/16\"", "PTR ASTM A36 1\" x 1\" x C.11",
                "REDONDO AISI 1018 Ø 1 1/2\"", "REDONDO AISI 1018 Ø 1 1/4\"", "REDONDO AISI 1018 Ø 1 3/8\"", "REDONDO AISI 1018 Ø 1\"", "REDONDO AISI 1018 Ø 1/2\"", "REDONDO AISI 1018 Ø 2 1/2\"", "REDONDO AISI 1018 Ø 2 5/8\"", "REDONDO AISI 1018 Ø 2\"", "REDONDO AISI 1018 Ø 3\"", "REDONDO AISI 1018 Ø 3/4\"", "REDONDO AISI 1018 Ø 7/8\"", "REDONDO NEGRO Ø 5/16\"", "REDONDO NEGRO Ø 5/8\"",
                "RIEL DE ACERO A36 1500", "SOLERA ASTM A36 1 1/2\" x 1/2\"", "SOLERA ASTM A36 1 1/4\" x 1/4\"", "SOLERA ASTM A36 1\" x 1/2\"", "SOLERA ASTM A36 2\" x 1\"", "SOLERA ASTM A36 4\" x 1\"", "SOLERA ASTM A36 6\" x 1\"", "SOLERA DE ALUMINIO ASTM A36 2\" X 1\"",
                "TOLDO ALUMINIO C.19",
                "TUBO DE ACERO A500 °B Ø 1 1/2\" CED. 80", "TUBO DE ACERO A500 °B Ø 1 1/2\" CED. 80 SIN/COS", "TUBO DE ACERO A500 °B Ø 1\" CED. 40 C/COS", "TUBO DE ACERO A500 °B Ø 1\" CED. 40 SIN/COS",
                "TUBO DE ALUMINIO B241 Ø  2\" x  1\" x 1/8\"", "TUBO DE ALUMINIO B241 Ø 2 1/2\"", "TUBO DE ALUMINIO B241 Ø 2\"", "TUBO DE ALUMINIO B241 Ø 2\" x  1\" x 1/8\"", "TUBO DE ALUMINIO B241 Ø 3 1/2\"", "TUBO DE ALUMINIO NEGRO B241 Ø 2 1/2\"",
                "TUBO STROCK CROMADO ASTM 1045 Ø 70mm X 63mm",
                "PERFIL ALUMINIO CUERNO EA 685 6061T6", "PERFIL ALUMINIO PRINCIPAL EXT 684 6061T6", "PERFIL ALUMINIO ANGULO VISTA 686", "PERFIL ALUMINIO REFUERZO INT 683", "BORDA LATERAL BASCULANTE 4.9 6061", "BORDA LATERAL BASCULANTE 3.5 6061", "PERFIL DE ALUMINIO TIPO BISAGRA ABATIBLE 3.10 MT 6061-T6", "PERFIL DE ALUMINIO TIPO ESCALON ABATIBLE 3.10 MT 6061-T6"
            ]
            
            for mat in materiales_iniciales:
                cursor.execute("IF NOT EXISTS (SELECT * FROM Tbl_Materiales_Aprobados WHERE Material = ?) INSERT INTO Tbl_Materiales_Aprobados (Material) VALUES (?)", (mat, mat))
            
            conn.commit()
            print(f"--- TB_MATERIALES_APROBADOS INICIALIZADA ({len(materiales_iniciales)} items) ---")

        # --- TABLAS DE JERARQUÍA DE PROYECTOS ---
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Proyectos_Tracto')
            BEGIN
                CREATE TABLE Tbl_Proyectos_Tracto (
                    ID_Tracto INT IDENTITY(1,1) PRIMARY KEY,
                    Nombre_Tracto VARCHAR(200) UNIQUE NOT NULL
                );
            END
        """)
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Tipos_Proyecto')
            BEGIN
                CREATE TABLE Tbl_Tipos_Proyecto (
                    ID_Tipo INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Tracto INT NOT NULL,
                    Nombre_Tipo VARCHAR(200) NOT NULL,
                    CONSTRAINT FK_Tipo_Tracto FOREIGN KEY (ID_Tracto) REFERENCES Tbl_Proyectos_Tracto(ID_Tracto) ON DELETE CASCADE
                );
            END
        """)
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Versiones_Ingenieria')
            BEGIN
                CREATE TABLE Tbl_Versiones_Ingenieria (
                    ID_Version INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Tipo INT NOT NULL,
                    Nombre_Version VARCHAR(200) NOT NULL,
                    CONSTRAINT FK_Version_Tipo FOREIGN KEY (ID_Tipo) REFERENCES Tbl_Tipos_Proyecto(ID_Tipo) ON DELETE CASCADE
                );
            END
        """)
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Clientes_Configuracion')
            BEGIN
                CREATE TABLE Tbl_Clientes_Configuracion (
                    ID_Config_Cliente INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Version INT NOT NULL,
                    Nombre_Cliente VARCHAR(200) NOT NULL,
                    CONSTRAINT FK_Cliente_Version FOREIGN KEY (ID_Version) REFERENCES Tbl_Versiones_Ingenieria(ID_Version) ON DELETE CASCADE
                );
            END
        """)
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_BOM_Revisiones')
            BEGIN
                CREATE TABLE Tbl_BOM_Revisiones (
                    ID_Revision INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Version INT NOT NULL,
                    Numero_Revision INT NOT NULL,
                    Estado VARCHAR(200) NOT NULL,
                    Fecha_Creacion DATETIME DEFAULT GETDATE(),
                    CONSTRAINT FK_Revision_Version2 FOREIGN KEY (ID_Version) REFERENCES Tbl_Versiones_Ingenieria(ID_Version) ON DELETE CASCADE
                );
            END
        """)
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Estaciones')
            BEGIN
                CREATE TABLE Tbl_Estaciones (
                    ID_Estacion INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Revision INT NOT NULL,
                    Nombre_Estacion VARCHAR(200) NOT NULL,
                    Orden INT NOT NULL DEFAULT 0,
                    CONSTRAINT FK_Estacion_Revision FOREIGN KEY (ID_Revision) REFERENCES Tbl_BOM_Revisiones(ID_Revision) ON DELETE CASCADE
                );
            END
        """)
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Ensambles')
            BEGIN
                CREATE TABLE Tbl_Ensambles (
                    ID_Ensamble INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Estacion INT NOT NULL,
                    Nombre_Ensamble VARCHAR(200) NOT NULL,
                    CONSTRAINT FK_Ensamble_Estacion FOREIGN KEY (ID_Estacion) REFERENCES Tbl_Estaciones(ID_Estacion) ON DELETE CASCADE
                );
            END
        """)
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_BOM_Estructura')
            BEGIN
                CREATE TABLE Tbl_BOM_Estructura (
                    ID_BOM INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Ensamble INT NOT NULL,
                    Codigo_Pieza VARCHAR(50) NOT NULL,
                    Cantidad FLOAT NOT NULL,
                    Observaciones VARCHAR(500),
                    CONSTRAINT FK_BOM_Ensamble FOREIGN KEY (ID_Ensamble) REFERENCES Tbl_Ensambles(ID_Ensamble) ON DELETE CASCADE
                );
            END
        """)
        
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'Tbl_Unidades_Fisicas')
            BEGIN
                CREATE TABLE Tbl_Unidades_Fisicas (
                    ID_Unidad INT IDENTITY(1,1) PRIMARY KEY,
                    ID_Revision INT NOT NULL,
                    VIN VARCHAR(50) NOT NULL,
                    Notas VARCHAR(MAX),
                    CONSTRAINT FK_Unidad_Revision FOREIGN KEY (ID_Revision) REFERENCES Tbl_BOM_Revisiones(ID_Revision) ON DELETE CASCADE
                );
            END
        """)
        conn.commit()

        print("--- ✅ SISTEMA DE AUDITORIA INICIALIZADO CORRECTAMENTE ---")
    except Exception as e:
        print(f"--- ⚠️ ALERTA SQL (Auditoría): {e} ---")
    finally:
        try:
             conn.close()
        except:
             pass

def registrar_auditoria(cursor, codigo_pieza, accion, valor_anterior, valor_nuevo, usuario):
    """
    Registra un evento en Tbl_Auditoria_Cambios.
    Maneja la conversión de dicts a JSON string si es necesario.
    """
    try:
        # Convertir a cadena si son diccionarios/listas
        if isinstance(valor_anterior, (dict, list)):
            valor_anterior = str(valor_anterior) # Usamos str() para ser consistente con lo que espera el parser (ast.literal_eval/json)
        if isinstance(valor_nuevo, (dict, list)):
            valor_nuevo = str(valor_nuevo)

        cursor.execute("""
            INSERT INTO Tbl_Auditoria_Cambios (Codigo_Pieza, Accion, Valor_Anterior, Valor_Nuevo, Usuario, Fecha_Hora)
            VALUES (?, ?, ?, ?, ?, GETDATE())
        """, (codigo_pieza, accion, str(valor_anterior), str(valor_nuevo), usuario))
    except Exception as e:
        print(f"--- ⚠️ ERROR AUDITORIA INTERNA: {e} ---")

@app.get("/api/catalog")
async def get_catalog():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        query = "SELECT * FROM Tbl_Maestro_Piezas"
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        data = []
        for row in cursor.fetchall():
            record = {}
            for col, val in zip(columns, row):
                record[col] = val if val is not None else "-"
            data.append(record)
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# 6. EDICIÓN DE MATERIALES (ESPECÍFICA + CAMPOS NUEVOS + PROCESO 3)
@app.put("/api/material/update")
async def update_material(request: Request, payload: Dict[str, Any]):
    print(f"--- UPDATE MATERIAL FULL EDITOR V2 ---")
    print(f"Payload: {payload}")

    # Extraer ID
    codigo_pieza = payload.get('Codigo_Pieza')
    codigo_legacy = payload.get('Codigo')
    id_param = codigo_pieza if codigo_pieza else codigo_legacy

    if not id_param:
        raise HTTPException(status_code=400, detail="Falta Codigo_Pieza o Codigo")

    # Campos Permitidos (Whitelist) - Incluyendo Proceso_3
    allowed_fields = [
        'Descripcion', 'Medida', 'Material', 'Link_Drive', 
        'Simetria', 'Proceso_Primario', 'Proceso_1', 'Proceso_2', 'Proceso_3'
    ]
    
    # REGLA ESPEJO: Si está activa y se actualiza la descripción, también el material
    if REGLA_ESPEJO_ACTIVA and 'Descripcion' in payload:
        payload['Material'] = payload['Descripcion']

    usuario = payload.get('usuario') or payload.get('Modificado_Por') or 'Sistema'

    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Asegurar columna Modificado_Por
        try:
            cursor.execute("SELECT Modificado_Por FROM Tbl_Maestro_Piezas WHERE 1=0")
        except:
             conn.rollback()
             cursor.execute("ALTER TABLE Tbl_Maestro_Piezas ADD Modificado_Por NVARCHAR(50)")
             conn.commit()

        # Construcción Dinámica Segura de la Query
        set_clauses = []
        values = []
        
        for field in allowed_fields:
            if field in payload:
               set_clauses.append(f"{field} = ?")
               values.append(payload[field])
        
        if not set_clauses:
             return {"status": "ignored", "message": "No hay campos válidos para actualizar"}


        # Agregar Auditoría
        set_clauses.append("Modificado_Por = ?")
        values.append(usuario)
        
        set_clauses.append("Ultima_Actualizacion = GETDATE()")
        
        query_set = ", ".join(set_clauses)
        
        # --- AUDITORIA: CAPTURAR VALOR ANTERIOR ---
        try:
             # Seleccionamos todos los campos afectados + ID
             cols_to_select = ", ".join(allowed_fields)
             cursor.execute(f"SELECT {cols_to_select} FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = ?", (id_param,))
             row_prev = cursor.fetchone()
             
             valor_anterior = {}
             if row_prev:
                 for i, field in enumerate(allowed_fields):
                     valor_anterior[field] = str(row_prev[i]) if row_prev[i] is not None else ""
             else:
                 valor_anterior = "REGISTRO NO ENCONTRADO (Posible error en Update)"
        except Exception as audit_read_e:
             valor_anterior = f"ERROR LECTURA PREVIA: {audit_read_e}"
        # ------------------------------------------

        # Query Principal
        query = f"UPDATE Tbl_Maestro_Piezas SET {query_set} WHERE Codigo_Pieza = ?"
        values.append(id_param)
        
        print(f"SQL GENERADO: {query}")
        
        cursor.execute(query, values)
        
        if cursor.rowcount == 0:
            print("Fallback: Actualizando por Codigo...")
            query_fallback = f"UPDATE Tbl_Maestro_Piezas SET {query_set} WHERE Codigo = ?"
            cursor.execute(query_fallback, values)

        conn.commit()
        
        if cursor.rowcount > 0:
             # --- AUDITORIA: REGISTRAR CAMBIO ---
             registrar_auditoria(cursor, id_param, 'EDICION_CATALOGO', valor_anterior, payload, usuario)
             conn.commit() # Commit del log
             # -----------------------------------
             return {"status": "success", "message": "Actualizado correctamente"}
        else:
             raise HTTPException(status_code=404, detail="No se encontró registro (Codigo/Codigo_Pieza)")

    except Exception as e:
        conn.rollback()
        print(f"ERROR UPDATE: {e}")
        raise HTTPException(status_code=500, detail=f"SQL Error: {str(e)}")
    finally:
        conn.close()

# 5. MOTOR DE ARBITRAJE EXCEL
@app.post("/api/excel/procesar")
async def procesar_excel(file: UploadFile = File(...)):
    print(f"--- PROCESANDO BOM EXCEL: {file.filename} ---")
    contents = await file.read()
    
    try:
        # 1. Leer Excel (Memoria)
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        scan_data = []
        last_estacion = None
        last_ensamble = None
        
        # 2. Iterar filas (Start Row 6 - 0-indexed es 5, pero openpyxl es 1-based, así que min_row=6)
        start_row = 6
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # Mapeo por índice (0-based)
            # Col 3 (D): CODIGO_PIEZA
            # Col 4 (E): DESCRIPCION / MATERIAL
            # Col 5 (F): MEDIDA
            # Col 7 (H): SIMETRIA
            # Col 8 (I): PROCESO PRIMARIO
            # Col 9 (J): PROCESO 1
            # Col 10 (K): PROCESO 2
            # Col 11 (L): PROCESO 3
            
            if not row: continue

            # Forward Fill Logic
            estacion = row[1] if len(row) > 1 and row[1] is not None else last_estacion
            ensamble = row[2] if len(row) > 2 and row[2] is not None else last_ensamble
            
            if estacion: last_estacion = estacion
            if ensamble: last_ensamble = ensamble

            # Validar Codigo Pieza (Columna D - Index 3)
            if len(row) <= 3: continue
            raw_codigo = row[3]
            codigo_pieza = str(raw_codigo).strip() if raw_codigo else None
            
            if not codigo_pieza or codigo_pieza.lower() in ['none', 'codigo', 'codigo_pieza', '']:
                continue

            # Extracción segura con manejo de nulos
            def get_val(idx):
                if idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
                return ""

            scan_data.append({
                'Estacion': last_estacion,
                'Ensamble': last_ensamble,
                'Codigo_Pieza': codigo_pieza,
                'Cantidad': 0, 
                'Descripcion_Excel': get_val(4),
                'Medida_Excel': get_val(5),
                'Material_Excel': "", # No mapeado explícitamente en columna aparte, dejamos vacío (Regla Espejo lo llenará si aplica)
                'Simetria': get_val(7),
                'Proceso_Primario': get_val(8),
                'Proceso_1': get_val(9),
                'Proceso_2': get_val(10),
                'Proceso_3': get_val(11),
                'Link_Drive': "" # No mapeado en este bloque
            })

        # 3. Comparar contra SQL
        conn = get_db_connection()
        cursor = conn.cursor()
        
        conflictos = []
        
        for item in scan_data:
            cursor.execute("SELECT Descripcion, Medida, Material, Simetria, Proceso_Primario, Proceso_1, Proceso_2, Proceso_3, Link_Drive FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = ?", (item['Codigo_Pieza'],))
            row_sql = cursor.fetchone()
            
            status = "OK"
            detalles = []

            if not row_sql:
                status = "NUEVO"
                sql_data = {}
            else:
                desc_sql = (row_sql[0] or "").strip()
                med_sql = (row_sql[1] or "").strip()
                mat_sql = (row_sql[2] or "").strip()
                # Otros campos para mostrar en UI
                sim_sql = (row_sql[3] or "").strip()
                pp_sql = (row_sql[4] or "").strip()
                p1_sql = (row_sql[5] or "").strip()
                p2_sql = (row_sql[6] or "").strip()
                p3_sql = (row_sql[7] or "").strip()
                link_sql = (row_sql[8] or "").strip()

                sql_data = {
                    'Descripcion': desc_sql,
                    'Medida': med_sql,
                    'Material': mat_sql,
                    'Simetria': sim_sql,
                    'Proceso_Primario': pp_sql,
                    'Proceso_1': p1_sql,
                    'Proceso_2': p2_sql,
                    'Proceso_3': p3_sql,
                    'Link_Drive': link_sql
                }

                # Comparación Flexible (Case Insensitive)
                if item['Descripcion_Excel'].lower() != desc_sql.lower():
                     if item['Descripcion_Excel']: # Solo si excel tiene dato
                        status = "CONFLICTO"
                        detalles.append(f"Desc: '{item['Descripcion_Excel']}' vs SQL '{desc_sql}'")
                
                if item['Medida_Excel'].lower() != med_sql.lower():
                     if item['Medida_Excel']:
                        status = "CONFLICTO"
                        detalles.append(f"Med: '{item['Medida_Excel']}' vs SQL '{med_sql}'")

            if status != "OK":
                conflictos.append({
                    'Codigo_Pieza': item['Codigo_Pieza'],
                    'Estado': status,
                    'Detalles': "; ".join(detalles),
                    'Excel_Data': item,
                    'SQL_Data': sql_data # <--- DATOS FALTANTES
                })

        return {
            "total_leidos": len(scan_data),
            "conflictos": conflictos,
            "mensaje": f"Procesado exitoso. {len(conflictos)} conflictos detectados."
        }

    except Exception as e:
        print(f"ERROR EXCEL: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando Excel: {str(e)}")

class SincronizacionItem(BaseModel):
    Codigo_Pieza: str
    Descripcion: Optional[str] = None
    Medida: Optional[str] = None
    Material: Optional[str] = None
    Link_Drive: Optional[str] = None
    Simetria: Optional[str] = None
    Proceso_Primario: Optional[str] = None
    Proceso_1: Optional[str] = None
    Proceso_2: Optional[str] = None
    Proceso_3: Optional[str] = None
    Modificado_Por: Optional[str] = None 
    Estado: str 
    
    model_config = ConfigDict(extra='ignore')

@app.post("/api/excel/sincronizar")
async def sincronizar_excel(items: List[SincronizacionItem]):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    procesados = 0
    errores = 0
    
    try:
        for item in items:
            # REGLA ESPEJO
            if REGLA_ESPEJO_ACTIVA:
                item.Material = item.Descripcion

            # Sanitizar datos (Evitar NULLs -> Strings Vacíos)
            desc = item.Descripcion if item.Descripcion is not None else ""
            medida = item.Medida if item.Medida is not None else ""
            material = item.Material if item.Material is not None else ""
            link = item.Link_Drive if item.Link_Drive is not None else ""
            simetria = item.Simetria if item.Simetria is not None else ""
            proc_prim = item.Proceso_Primario if item.Proceso_Primario is not None else ""
            proc_1 = item.Proceso_1 if item.Proceso_1 is not None else ""
            proc_2 = item.Proceso_2 if item.Proceso_2 is not None else ""
            proc_3 = item.Proceso_3 if item.Proceso_3 is not None else ""
            
            # Auditoría
            usuario = item.Modificado_Por if item.Modificado_Por else "Importador Excel"

            if item.Estado == "NUEVO":
                # Lógica de Inserción (INSERT COMPLETO)
                cursor.execute("""
                    IF NOT EXISTS (SELECT 1 FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = ?)
                    BEGIN
                        INSERT INTO Tbl_Maestro_Piezas 
                        (Codigo_Pieza, Descripcion, Medida, Material, Simetria, Proceso_Primario, Proceso_1, Proceso_2, Proceso_3, Link_Drive, Ultima_Actualizacion, Modificado_Por)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), ?)
                    END
                """, (item.Codigo_Pieza, item.Codigo_Pieza, desc, medida, material, simetria, proc_prim, proc_1, proc_2, proc_3, link, usuario))
                

                if cursor.rowcount > 0:
                    procesados += 1
                    # Log Auditoría CREACIÓN
                    registrar_auditoria(cursor, item.Codigo_Pieza, 'CREACION', 'NO EXISTIA', item.model_dump(), usuario)

            elif item.Estado == "CONFLICTO":
                # Lógica de Actualización (UPDATE COMPLETO)
                # 1. Obtener datos anteriores
                cursor.execute("SELECT * FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = ?", (item.Codigo_Pieza,))
                row_old = cursor.fetchone()
                val_anterior = str(row_old) if row_old else "DESCONOCIDO"

                # 2. Ejecutar Update
                cursor.execute("""
                    UPDATE Tbl_Maestro_Piezas
                    SET Descripcion = ?,
                        Medida = ?,
                        Material = ?,
                        Simetria = ?,
                        Proceso_Primario = ?,
                        Proceso_1 = ?,
                        Proceso_2 = ?,
                        Proceso_3 = ?,
                        Link_Drive = ?,
                        Ultima_Actualizacion = GETDATE(),
                        Modificado_Por = ?
                    WHERE Codigo_Pieza = ?
                """, (desc, medida, material, simetria, proc_prim, proc_1, proc_2, proc_3, link, usuario, item.Codigo_Pieza))
                
                if cursor.rowcount > 0:
                    procesados += 1
                    # Log Auditoría MODIFICACIÓN
                    registrar_auditoria(cursor, item.Codigo_Pieza, 'MODIFICACION', val_anterior, item.model_dump(), usuario)

        conn.commit()
        return {"status": "ok", "message": f"{procesados} registros sincronizados exitosamente."}

    except Exception as e:
        conn.rollback()
        print(f"Error en sincronizacion: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al sincronizar BD: {str(e)}")
    finally:
        conn.close()

# 7. CONFIGURACIÓN Y UTILIDADES (ACTUALIZADOR DE LINKS)
@app.post("/api/config/update_links")
async def update_links(payload: Dict[str, str]):
    root_path = payload.get('root_path')
    if not root_path or not os.path.exists(root_path):
        raise HTTPException(status_code=400, detail="Ruta base inválida o inaccesible")

    # Archivo Maestro definido por el usuario
    excel_path = os.path.join(root_path, "MAESTRO DE MATERIALES.xlsx")
    if not os.path.exists(excel_path):
        print(f"ERROR: No se encontró {excel_path}")
        # Retornamos error claro para el frontend
        raise HTTPException(status_code=404, detail=f"No se encontró el archivo 'MAESTRO DE MATERIALES.xlsx' en {root_path}")

    print(f"--- SINCRONIZANDO ENLACES DESDE EXCEL: {excel_path} ---")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # 1. Leer Excel usando pandas (según imagen: Col A=Codigo, Col B=URL_Google_Drive)
        df = pd.read_excel(excel_path)
        
        # Normalizar nombres de columnas
        df.columns = [str(c).strip() for c in df.columns]
        
        # Validar Columnas (Basado en captura de pantalla)
        if 'Codigo' not in df.columns:
            raise HTTPException(status_code=400, detail="El Excel no tiene la columna 'Codigo'")
        
        # Buscar columna de Drive (puede ser 'URL_Google_Drive' o similar)
        drive_col = next((c for c in df.columns if 'drive' in c.lower() or 'url' in c.lower()), None)
        
        if not drive_col:
             raise HTTPException(status_code=400, detail="No se encontró la columna de enlaces de Drive")

        updated_count = 0
        
        # 2. Iterar y Actualizar
        for _, row in df.iterrows():
            codigo = str(row['Codigo']).strip()
            link = str(row[drive_col]).strip()
            

            # Solo actualizar si el link existe y no es nulo
            if codigo and link and link.lower() != 'nan' and link != "":
                
                # --- AUDITORIA: CAPTURAR LINK ANTERIOR ---
                prev_link_val = "N/A"
                try:
                    cursor.execute("SELECT Link_Drive FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = ?", (codigo,))
                    row_link = cursor.fetchone()
                    if row_link:
                         prev_link_val = row_link[0]
                except:
                    pass
                # ----------------------------------------

                cursor.execute("""
                    UPDATE Tbl_Maestro_Piezas 
                    SET Link_Drive = ?, 
                        Ultima_Actualizacion = GETDATE(),
                        Modificado_Por = 'Sincronizador Excel'
                    WHERE Codigo_Pieza = ?
                """, (link, codigo))
                
                if cursor.rowcount > 0:
                     updated_count += cursor.rowcount
                     # --- AUDITORIA ---
                     registrar_auditoria(cursor, codigo, 'ACTUALIZACION_LINKS', prev_link_val, link, 'Sincronizador Excel')
                     # -----------------

        conn.commit()
        print(f"--- SINCRONIZACIÓN EXCEL FINALIZADA: {updated_count} links actualizados ---")
        return {"status": "ok", "updated": updated_count}

    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        print(f"ERROR EN SINCRONIZACIÓN EXCEL: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.post("/api/excel/actualizar_enlaces")
async def actualizar_enlaces_manual(file: UploadFile = File(...)):
    """
    Actualiza enlaces Drive desde un Excel cargado por el usuario.
    Estructura: Col 0 = Código, Col 1 = Link_Drive
    """
    print(f"--- ACTUALIZANDO ENLACES DESDE EXCEL MANUAL: {file.filename} ---")
    contents = await file.read()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # 1. Leer Excel (Memoria)
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        updated_count = 0
        row_idx = 0
        
        # 2. Iterar filas
        for row in ws.iter_rows(values_only=True):
            row_idx += 1
            # Saltar encabezado (fila 1)
            if row_idx == 1:
                continue
                
            if not row or len(row) < 2:
                continue
                
            codigo = str(row[0]).strip() if row[0] else None
            link = str(row[1]).strip() if row[1] else None
            

            # Solo procesar si hay código y link válido
            if codigo and link and link.lower() != 'nan' and link != "" and link != "-":
                
                # --- AUDITORIA: CAPTURAR LINK ANTERIOR ---
                prev_link_val = "N/A"
                try:
                    cursor.execute("SELECT Link_Drive FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = ?", (codigo,))
                    row_link = cursor.fetchone()
                    if row_link:
                         prev_link_val = row_link[0]
                except:
                    pass
                # ----------------------------------------

                cursor.execute("""
                    UPDATE Tbl_Maestro_Piezas 
                    SET Link_Drive = ?, 
                        Ultima_Actualizacion = GETDATE(),
                        Modificado_Por = 'Sincronizador Manual (Excel)'
                    WHERE Codigo_Pieza = ?
                """, (link, codigo))
                
                if cursor.rowcount == 0:
                    # Fallback eliminado: La columna 'Codigo' no existe en esta versión de la BD.
                    # Si se requiere soporte legacy, asegurar que la columna exista primero.
                    print(f"--- AVISO: Codigo '{codigo}' no encontrado por Codigo_Pieza ---")
                
                else: 
                     updated_count += cursor.rowcount
                     # --- AUDITORIA ---
                     registrar_auditoria(cursor, codigo, 'ACTUALIZACION_LINKS', prev_link_val, link, 'Sincronizador Manual (Excel)')
                     # -----------------

        conn.commit()
        print(f"--- ACTUALIZACIÓN MANUAL FINALIZADA: {updated_count} enlaces actualizados ---")
        return {"status": "ok", "actualizados": updated_count}

    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        print(f"ERROR EN ACTUALIZACIÓN MANUAL: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando Excel: {str(e)}")
    finally:
        if 'conn' in locals(): conn.close()

# --- FASE 12 y 13: AUDITOR AVANZADO Y HERRAMIENTAS ---

@app.post("/api/excel/auditar")
async def auditar_excel(file: UploadFile = File(...)):
    """
    Audita un archivo Excel comparando múltiples columnas con la BD.
    Retorna errores puntuales para UI y reporte detallado para Excel.
    """
    print(f"--- INICIANDO AUDITORÍA AVANZADA: {file.filename} ---")
    contents = await file.read()
    
    errores = []
    reporte_detallado = [] # Lista de objetos con contexto completo
    
    field_map = {
        'Descripcion': 4,
        'Medida': 5,
        'Simetria': 7,
        'Proceso_Primario': 8,
        'Proceso_1': 9,
        'Proceso_2': 10,
        'Proceso_3': 11
    }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
            if not row or len(row) < 12: 
                continue
            
            codigo_excel = str(row[3]).strip() if row[3] else None
            if not codigo_excel: continue

            cursor.execute("""
                SELECT Descripcion, Medida, Simetria, Proceso_Primario, Proceso_1, Proceso_2, Proceso_3 
                FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = ?
            """, (codigo_excel,))
            row_bd = cursor.fetchone()
            
            if row_bd:
                vals_bd = {
                    'Descripcion': str(row_bd[0] or "").strip(),
                    'Medida': str(row_bd[1] or "").strip(),
                    'Simetria': str(row_bd[2] or "").strip(),
                    'Proceso_Primario': str(row_bd[3] or "").strip(),
                    'Proceso_1': str(row_bd[4] or "").strip(),
                    'Proceso_2': str(row_bd[5] or "").strip(),
                    'Proceso_3': str(row_bd[6] or "").strip(),
                }
                
                vals_excel = {}
                row_diffs = []

                # Recolectar datos y diffs
                for field, col_idx in field_map.items():
                    val_excel = str(row[col_idx]).strip() if row[col_idx] else ""
                    vals_excel[field] = val_excel
                    
                    if val_excel != vals_bd[field]:
                         errores.append({
                            "fila": row_idx,
                            "codigo": codigo_excel,
                            "campo": field,
                            "excel": val_excel,
                            "bd": vals_bd[field]
                        })
                         row_diffs.append(field)
                
                # Si hubo diferencias en esta fila, guardamos contexto completo
                if row_diffs:
                    reporte_detallado.append({
                        "fila": row_idx,
                        "codigo": codigo_excel,
                        "excel_data": vals_excel,
                        "bd_data": vals_bd,
                        "campos_error": row_diffs
                    })

        print(f"--- AUDITORÍA FINALIZADA: {len(errores)} discrepancias en {len(reporte_detallado)} filas ---")
        return {"status": "ok", "errores": errores, "reporte_detallado": reporte_detallado}

    except Exception as e:
        print(f"ERROR AUDITORIA: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

from datetime import datetime
import json

@app.post("/api/excel/corregir")
async def corregir_excel(
    file: UploadFile = File(...), 
    correcciones: str = Form(...) # JSON String
):
    print(f"--- INICIANDO AUTOCORRECCIÓN SEGURA: {file.filename} ---")

    try:
        corrections_list = json.loads(correcciones)
        
        # Leer archivo en memoria
        contents = await file.read()
        
        # 2. APLICAR CORRECCIONES
        wb = openpyxl.load_workbook(io.BytesIO(contents)) # No data_only para preservar FÓRMULAS
        ws = wb.active # Asumimos hoja activa

        
        # Mapeo Campo -> Columna (1-based para cell.column)
        # D(4)=Codigo, E(5)=Desc, F(6)=Medida, H(8)=Simetria
        # PROCESOS DISTRIBUIDOS (NO COMBINADOS):
        # I(9)=Primario, J(10)=Proc1, K(11)=Proc2, L(12)=Proc3
        col_map = {
            'Descripcion': 5, # E
            'Medida': 6,      # F
            'Simetria': 8,    # H
            'Proceso_Primario': 9, # I
            'Proceso_1': 10,  # J
            'Proceso_2': 11,  # K
            'Proceso_3': 12   # L
        }

        count = 0
        for item in corrections_list:
            fila = int(item['fila'])
            campo = item['campo']
            valor_correcto = item['bd']
            
            if campo in col_map:
                col_idx = col_map[campo]
                # openpyxl: ws.cell(row=X, column=Y).value = ...
                ws.cell(row=fila, column=col_idx).value = valor_correcto
                count += 1
        
        # 3. GUARDAR COMO BINARIO Y DEVOLVER
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"Archivo corregido en memoria ({count} cambios). Enviando al cliente...")
        
        headers = {
            'Content-Disposition': f'attachment; filename="CORREGIDO_{file.filename}"'
        }
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers=headers
        )

    except Exception as e:
        print(f"ERROR CORRECCIÓN: {e}")
        raise HTTPException(status_code=500, detail=f"Error al corregir archivo: {str(e)}")

@app.post("/api/system/open_file")
async def open_file_endpoint(payload: Dict[str, str]):
    path = payload.get('path')
    if not path or not os.path.exists(path):
         raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    try:
        os.startfile(path)
        return {"status": "ok"}
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/exportar_reporte")
async def exportar_reporte(payload: List[Dict[str, Any]]):
    """
    Genera reporte estilo comparativo:
    Fila Excel
    Fila BD (Errors Highlighted)
    [Empty Row]
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Auditoria"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Rojo claro
        bd_row_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid") # Gris muy claro
        
        headers = ['Fila', 'Código', 'Fuente', 'Descripción', 'Medida', 'Simetría', 'Proceso Primario', 'Proceso 1', 'Proceso 2', 'Proceso 3']
        ws.append(headers)
        
        # Aplicar estilo headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        current_row = 2
        
        # Ordenar columnas para iteración
        col_keys = ['Descripcion', 'Medida', 'Simetria', 'Proceso_Primario', 'Proceso_1', 'Proceso_2', 'Proceso_3']

        for item in payload:
            fila_orig = item.get('fila', '-')
            codigo = item.get('codigo', '-')
            excel_data = item.get('excel_data', {})
            bd_data = item.get('bd_data', {})
            errores = item.get('campos_error', [])
            
            # --- FILA 1: EXCEL ---
            ws.cell(row=current_row, column=1, value=fila_orig)
            ws.cell(row=current_row, column=2, value=codigo)
            ws.cell(row=current_row, column=3, value="EXCEL").font = Font(bold=True)
            
            for idx, key in enumerate(col_keys, start=4):
                ws.cell(row=current_row, column=idx, value=excel_data.get(key, ""))
            
            # --- FILA 2: BASE DE DATOS ---
            next_row = current_row + 1
            ws.cell(row=next_row, column=1, value=fila_orig)
            ws.cell(row=next_row, column=2, value=codigo)
            ws.cell(row=next_row, column=3, value="BASE DATOS").font = Font(bold=True)
            
            for idx, key in enumerate(col_keys, start=4):
                cell = ws.cell(row=next_row, column=idx, value=bd_data.get(key, ""))
                cell.fill = bd_row_fill # Default BD style
                
                # Highlight si hay error
                if key in errores:
                    cell.fill = error_fill
                    cell.font = Font(bold=True, color="CC0000")

            # Separador (Row vacía)
            current_row += 3 

        # Auto-width básico
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="Reporte_Auditoria_Avanzado.xlsx"'
        }
        return Response(content=output.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    except Exception as e:
        print(f"ERROR REPORTE: {e}")
        raise HTTPException(status_code=500, detail=str(e))

import ast

@app.get("/api/historial")
async def obtener_historial(busqueda: Optional[str] = None, limite: int = 50):
    print(f"--- CONSULTANDO HISTORIAL (Busqueda: {busqueda}, Limite: {limite}) ---")
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT TOP (?) ID_Log, Codigo_Pieza, Accion, Valor_Anterior, Valor_Nuevo, Usuario, Fecha_Hora 
            FROM Tbl_Auditoria_Cambios 
        """
        params = [limite]
        
        if busqueda:
            query += " WHERE Codigo_Pieza LIKE ? OR Usuario LIKE ? "
            search_term = f"%{busqueda}%"
            params.extend([search_term, search_term])
            
        query += " ORDER BY Fecha_Hora DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        historial = []
        for row in rows:
            val_ant = row[3]
            val_nue = row[4]
            
            # Intentar parsear Valor_Anterior
            if val_ant and isinstance(val_ant, str):
                try:
                    if val_ant.strip().startswith('{') or val_ant.strip().startswith('['):
                         val_ant = json.loads(val_ant.replace("'", '"')) # Attempt JSON fix or standard load
                    elif val_ant.strip().startswith('('):
                         val_ant = ast.literal_eval(val_ant)
                    else:
                         # Try literal eval for python dict repr
                         val_ant = ast.literal_eval(val_ant)
                except:
                    pass # Keep as string if fail

            # Intentar parsear Valor_Nuevo
            if val_nue and isinstance(val_nue, str):
                try:
                    if val_nue.strip().startswith('{') or val_nue.strip().startswith('['):
                         val_nue = json.loads(val_nue.replace("'", '"'))
                    elif val_nue.strip().startswith('('):
                         val_nue = ast.literal_eval(val_nue)
                    else:
                         val_nue = ast.literal_eval(val_nue)
                except:
                    pass

            historial.append({
                "id": row[0],
                "codigo": row[1],
                "accion": row[2],
                "valor_anterior": val_ant,
                "valor_nuevo": val_nue,
                "usuario": row[5],
                "fecha": row[6].strftime("%Y-%m-%d %H:%M:%S") if row[6] else None
            })
            
        return historial

    except Exception as e:
        print(f"ERROR HISTORIAL: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


# --- FASE 18: ESTANDARIZACIÓN DE DATOS ---

@app.get("/api/limpieza/descripciones_unicas")
async def get_unique_descriptions():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        query = """
            SELECT Descripcion, COUNT(Codigo_Pieza) as Total 
            FROM Tbl_Maestro_Piezas 
            WHERE Descripcion IS NOT NULL AND Descripcion != ''
            GROUP BY Descripcion 
            ORDER BY Descripcion ASC
        """
        cursor.execute(query)
        data = [{"descripcion": row[0], "total": row[1]} for row in cursor.fetchall()]
        return data
    except Exception as e:
         print(f"ERROR DESC UNICAS: {e}")
         raise HTTPException(status_code=500, detail=str(e))
    finally:
         conn.close()

class MasivoUpdate(BaseModel):
    old_desc: str
    new_desc: str
    usuario: str

@app.post("/api/limpieza/actualizar_masivo")
async def actualizar_masivo(payload: MasivoUpdate):
    print(f"--- INICIANDO ESTANDARIZACION MASIVA: '{payload.old_desc}' -> '{payload.new_desc}' ---")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # 1. Obtener todas las piezas afectadas para auditoría individual
        cursor.execute("SELECT Codigo_Pieza FROM Tbl_Maestro_Piezas WHERE Descripcion = ?", (payload.old_desc,))
        piezas = cursor.fetchall()
        
        if not piezas:
             return {"status": "ignored", "message": "No se encontraron piezas con esa descripción."}
        
        actualizadas = 0
        
        # 2. Iterar y actualizar UNO A UNO
        for p in piezas:
            codigo = p[0]
            
            val_nuevo = ""
            
            # REGLA ESPEJO
            if globals().get('REGLA_ESPEJO_ACTIVA', True):
                cursor.execute("""
                    UPDATE Tbl_Maestro_Piezas 
                    SET Descripcion = ?, Material = ?, Ultima_Actualizacion = GETDATE(), Modificado_Por = ?
                    WHERE Codigo_Pieza = ?
                """, (payload.new_desc, payload.new_desc, payload.usuario, codigo))
                val_nuevo = str({"Descripcion": payload.new_desc, "Material": payload.new_desc})
            else:
                cursor.execute("""
                    UPDATE Tbl_Maestro_Piezas 
                    SET Descripcion = ?, Ultima_Actualizacion = GETDATE(), Modificado_Por = ?
                    WHERE Codigo_Pieza = ?
                """, (payload.new_desc, payload.usuario, codigo))
                val_nuevo = str({"Descripcion": payload.new_desc})
            
            if cursor.rowcount > 0:
                actualizadas += 1
                # Auditoría Individual
                registrar_auditoria(
                    cursor, 
                    codigo_pieza=codigo, 
                    accion='ESTANDARIZACION_MASIVA', 
                    valor_anterior=str({'Descripcion': payload.old_desc}), 
                    valor_nuevo=val_nuevo, 
                    usuario=payload.usuario
                )
        
        conn.commit()
        print(f"--- ESTANDARIZACION FINALIZADA: {actualizadas} piezas actualizadas ---")
        return {"status": "ok", "actualizadas": actualizadas}
        
    except Exception as e:
        conn.rollback()
        print(f"ERROR MASIVO: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8001)
