import sys
import io
import json
import os
import pyodbc
import pandas as pd
from sqlalchemy import create_engine, text
import urllib.parse
import datetime
import platform
import base64
import difflib
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import time

PATH_MAP_FILE = "file_paths_map.json"

def ensure_v13_1_schema():
    engine = get_engine()
    with engine.begin() as conn:
        # 1. Tabla de Fuentes de Datos
        conn.execute(text("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Tbl_Fuentes_Datos' AND xtype='U')
            BEGIN
                CREATE TABLE Tbl_Fuentes_Datos (
                    ID INT IDENTITY(1,1) PRIMARY KEY,
                    Nombre_Logico NVARCHAR(255),
                    Ruta_Actual NVARCHAR(MAX),
                    Ultima_Sincronizacion DATETIME,
                    Estado NVARCHAR(50) DEFAULT 'ACTIVO'
                )
            END
        """))
        
        # 2. Columna Simetria en Maestro
        try:
            conn.execute(text("ALTER TABLE Tbl_Maestro_Piezas ADD Simetria NVARCHAR(100)"))
        except:
            pass # Ya existe
            
        # 3. Columnas de Auditoría Extendida
        try:
           conn.execute(text("ALTER TABLE Tbl_Auditoria_Conflictos ADD Simetria_Excel NVARCHAR(100)"))
           conn.execute(text("ALTER TABLE Tbl_Auditoria_Conflictos ADD Proceso_Primario_Excel NVARCHAR(100)"))
           conn.execute(text("ALTER TABLE Tbl_Auditoria_Conflictos ADD Tipo_Conflicto NVARCHAR(50)"))
        except:
            pass

def get_sources():
    ensure_v13_1_schema()
    engine = get_engine()
    with engine.connect() as conn:
        return sanitize(pd.read_sql("SELECT * FROM Tbl_Fuentes_Datos WHERE Estado = 'ACTIVO'", conn)).to_dict(orient='records')

def add_source(name, path):
    ensure_v13_1_schema()
    engine = get_engine()
    with engine.begin() as conn:
        conn.execute(text("INSERT INTO Tbl_Fuentes_Datos (Nombre_Logico, Ruta_Actual) VALUES (:n, :r)"), {"n": name, "r": path})
    return {"status": "success"}

def register_path(name, path):
    """Alias para mantener compatibilidad, llama al nuevo método add_source"""
    return add_source(name, path)

def update_source(id, path):
    engine = get_engine()
    with engine.begin() as conn:
        conn.execute(text("UPDATE Tbl_Fuentes_Datos SET Ruta_Actual = :r WHERE ID = :id"), {"r": path, "id": id})
    return {"status": "success"}

def scan_and_ingest(source_id):
    engine = get_engine()
    
    # 1. Obtener Ruta
    path = None
    with engine.connect() as conn:
        res = conn.execute(text("SELECT Ruta_Actual FROM Tbl_Fuentes_Datos WHERE ID = :id"), {"id": source_id}).fetchone()
        if res:
            path = res[0]
        else:
            return {"status": "error", "message": "Origen de datos no encontrado en la Base de Datos. Intente recargar la lista."}
        
    if not path or not os.path.exists(path):
        return {"status": "error", "message": f"Archivo no encontrado o ruta inválida: {path}"}
        
    # 2. Leer Excel (Data Only)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active # Asumimos hoja activa o primera
        
        # Mapeo Columnas (0-based en iter_rows, pero Excel es 1-based)
        # D=4, E=5, F=6, H=8, I=9, J-L=10-12
        # openpyxl col idx: A=1, B=2 ... D=4
        
        new_count = 0
        conflict_count = 0
        
        with engine.begin() as conn:
            # Cache de códigos existentes para evitar miles de SELECTs
            existing_codes = set(pd.read_sql("SELECT Codigo_Pieza FROM Tbl_Maestro_Piezas", conn)['Codigo_Pieza'].str.upper().tolist())
            
            for row in ws.iter_rows(min_row=6, values_only=True):
                # row es tupla (0..N). Index 3 es Col D (Codigo).
                if not row or len(row) < 5: continue
                
                code = str(row[3]).strip().upper() if row[3] else ""
                if not code or code in ["NONE", "CODIGO", "CODE"]: continue
                
                # Extraer Data
                desc = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                medida = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                simetria = str(row[7]).strip() if len(row) > 7 and row[7] else "" # Col H (Index 7)
                proc_prim = str(row[8]).strip() if len(row) > 8 and row[8] else "" # Col I (Index 8)
                
                # Procesos Secundarios (Concat J, K, L)
                procs = []
                if len(row) > 9 and row[9]: procs.append(str(row[9]).strip())
                if len(row) > 10 and row[10]: procs.append(str(row[10]).strip())
                if len(row) > 11 and row[11]: procs.append(str(row[11]).strip())
                
                p1 = procs[0] if len(procs) > 0 else ""
                p2 = procs[1] if len(procs) > 1 else ""
                p3 = procs[2] if len(procs) > 2 else ""
                
                # CASO 1: NO EXISTE -> INSERTAR
                if code not in existing_codes:
                    conn.execute(text("""
                        INSERT INTO Tbl_Maestro_Piezas 
                        (Codigo_Pieza, Descripcion, Medida, Simetria, Proceso_Primario, Proceso_1, Proceso_2, Proceso_3, Ultima_Actualizacion)
                        VALUES (:c, :d, :m, :s, :p0, :p1, :p2, :p3, GETDATE())
                    """), {
                        'c': code, 'd': desc, 'm': medida, 's': simetria,
                        'p0': proc_prim, 'p1': p1, 'p2': p2, 'p3': p3
                    })
                    existing_codes.add(code)
                    new_count += 1
                else:
                    # CASO 2: YA EXISTE -> COMPARAR (Simplificado a Descripción para MVP, extendible)
                    # En v13.1 DATA MANAGER se pide auditar MULTIPLES columnas.
                    # Hacemos una validación rápida. Si difiere, lanzamos a conflicto.
                    
                    # Traer data actual (Optimizacion: Podriamos cachear todo, pero memoria...)
                    curr = conn.execute(text("SELECT Descripcion FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = :c"), {'c': code}).fetchone()
                    curr_desc = curr[0] if curr else ""
                    
                    if curr_desc != desc:
                        # Registrar Conflicto
                        # Upsert en Auditoria
                        conn.execute(text("""
                            MERGE Tbl_Auditoria_Conflictos AS target
                            USING (SELECT :code AS Codigo) AS source
                            ON (target.Codigo_Pieza = source.Codigo AND target.Estado = 'PENDIENTE')
                            WHEN MATCHED THEN
                                UPDATE SET Desc_Excel = :desc, Fecha_Deteccion = GETDATE()
                            WHEN NOT MATCHED THEN
                                INSERT (Codigo_Pieza, Desc_Excel, Desc_Master, Estado, Fecha_Deteccion, Tipo_Conflicto)
                                VALUES (:code, :desc, :master, 'PENDIENTE', GETDATE(), 'DATOS_DIFERENTES');
                        """), {'code': code, 'desc': desc, 'master': curr_desc})
                        conflict_count += 1

            # Actualizar timestamp fuente
            conn.execute(text("UPDATE Tbl_Fuentes_Datos SET Ultima_Sincronizacion = GETDATE() WHERE ID = :id"), {'id': source_id})
            
        return {"status": "success", "new_items": new_count, "conflicts": conflict_count}
        
    except Exception as e:
        return {"status": "error", "message": str(e)}


# FORZAR SALIDA UTF-8 (Vital para comunicación con Flutter)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# --- CONFIGURACIÓN ---
def load_config():
    config_path = "config.json"
    if hasattr(sys, '_MEIPASS'):
        config_path = os.path.join(os.path.dirname(sys.executable), "config.json")
    
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            return json.load(f)
    return {}

def get_base_path():
    if hasattr(sys, '_MEIPASS'):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_best_driver():
    """Selecciona el mejor driver ODBC disponible en el sistema."""
    try:
        drivers = pyodbc.drivers()
        # Lista de prioridad según recomendación de Microsoft y compatibilidad
        candidates = [
            'ODBC Driver 18 for SQL Server',
            'ODBC Driver 17 for SQL Server',
            'ODBC Driver 13 for SQL Server',
            'SQL Server Native Client 11.0',
            'SQL Server' # Fallback legacy (menos seguro, pero funcional)
        ]
        
        for candidate in candidates:
            if candidate in drivers:
                return candidate
                
        raise Exception("No se encontró ningún driver ODBC compatible (17, 18, 13, Native Client). Instale 'ODBC Driver 17 for SQL Server'.")
    except Exception as e:
        # En caso de error obteniendo lista, intentar forzar el 17 que es el estándar actual
        return 'ODBC Driver 17 for SQL Server'

def get_connection_string():
    config = load_config()
    server = config.get('server', '192.168.1.73,1433')
    database = config.get('database', 'DB_Materiales_Industrial')
    is_windows_auth = config.get('is_windows_auth', True)
    
    # Selección Dinámica del Driver (v13.3)
    driver = get_best_driver()
    
    # Parámetros adicionales según el driver
    # ODBC 18 requiere TrustServerCertificate=yes por defecto si no hay certificado válido
    extra_params = ""
    if '18' in driver or '17' in driver:
        extra_params = "TrustServerCertificate=yes;"
    
    if is_windows_auth:
        params = urllib.parse.quote_plus(
            f'DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection=yes;Encrypt=no;LoginTimeout=15;{extra_params}'
        )
    else:
        user = config.get('user', '')
        password = config.get('password', '')
        # Si trusted_connection es FALSE, usamos UID/PWD
        params = urllib.parse.quote_plus(
            f'DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={user};PWD={password};Encrypt=no;LoginTimeout=15;{extra_params}'
        )
    
    return f'mssql+pyodbc:///?odbc_connect={params}'

def save_sys_config(payload):
    try:
        config_path = "config.json"
        # Si estamos en frozen, guardar junto al ejecutable o en ruta estable
        if hasattr(sys, '_MEIPASS'):
             config_path = os.path.join(os.path.dirname(sys.executable), "config.json")
        
        current = load_config()
        
        # Merge safe keys
        valid_keys = ['server', 'database', 'user', 'password', 'blueprints_path', 'generics_path', 'trusted_connection']
        for k, v in payload.items():
            if k in valid_keys:
                current[k] = v
        
        # Logic for auth mode
        # Si 'trusted_connection' es explícito 'no' O hay user/pass, usar SQL Auth
        is_trusted = True
        if payload.get('trusted_connection') == 'no':
            is_trusted = False
        elif payload.get('user') and payload.get('password'):
            is_trusted = False
            
        current['is_windows_auth'] = is_trusted

        with open(config_path, 'w') as f:
            json.dump(current, f, indent=4)
            
        return {"status": "success", "message": "Configuración guardada"}
    except Exception as e:
        return {"status": "error", "message": f"Error guardando config: {str(e)}"}

def get_engine():
    return create_engine(get_connection_string())

def sanitize(df):
    # 1. Reemplazar NaN/None con cadena VACÍA "" (Para que Flutter muestre celda vacía, no "null")
    df = df.fillna("")
    
    # 2. Limpieza de Strings (Evitar crash de JSON por caracteres de Excel)
    # Convertir columnas de texto a string y limpiar saltos de línea/tabs que rompen JSON
    for col in df.select_dtypes(include=['object']):
        df[col] = df[col].astype(str).str.replace(r'[\r\n\t]+', ' ', regex=True)
        
    # 3. Fechas a string
    for col in df.select_dtypes(include=['datetime', 'datetimetz']).columns:
        df[col] = df[col].astype(str).replace('NaT', "")
        
    return df

# --- RUTAS DE API ---

def test_connection():
    try:
        engine = get_engine()
        driver_used = get_best_driver()
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return {"status": "success", "message": f"Conectado exitosamente usando: {driver_used}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def get_master_catalog():
    engine = get_engine()
    query = "SELECT * FROM Tbl_Maestro_Piezas ORDER BY Codigo_Pieza"
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)
    return sanitize(df).to_dict(orient='records')

def get_conflicts():
    engine = get_engine()
    query = "SELECT * FROM V_Auditoria_Conflictos"
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)
    
    # Aliases for frontend compatibility
    if not df.empty:
        df['id'] = df['Id'] if 'Id' in df.columns else ''
        df['codigo'] = df['Codigo_Pieza']
        df['descripcion'] = df['Descripcion_Final'] if 'Descripcion_Final' in df.columns else df['Desc_Master']
        df['archivo'] = df['Nombre_Archivo'] if 'Nombre_Archivo' in df.columns else ''
        df['hoja'] = df['Nombre_Hoja'] if 'Nombre_Hoja' in df.columns else ''
        df['fila'] = df['Numero_Fila_Excel'] if 'Numero_Fila_Excel' in df.columns else ''
        df['desc_excel'] = df['Desc_Excel'] if 'Desc_Excel' in df.columns else ''

    return sanitize(df).to_dict(orient='records')

def get_history(code):
    engine = get_engine()
    q = text("""
        SELECT TOP 50
            Codigo_Pieza as codigo, 
            Descripcion_Final as descripcion, 
            Estado_Resolucion as estado, 
            Fecha_Resolucion as fecha,
            Usuario as usuario
        FROM Tbl_Historial_Resoluciones 
        WHERE Codigo_Pieza = :c 
        ORDER BY Fecha_Resolucion DESC
    """)
    with engine.connect() as conn:
        df = pd.read_sql(q, conn, params={"c": code})
    return sanitize(df).to_dict(orient='records')

def get_resolved_tasks():
    engine = get_engine()
    query = """
    SELECT TOP 50
        Codigo_Pieza as codigo, 
        Descripcion_Final as descripcion, 
        Estado_Resolucion as estado, 
        Fecha_Resolucion as fecha,
        FORMAT(Fecha_Resolucion, 'dd/MM/yyyy HH:mm') as fecha_fmt,
        Usuario as usuario
    FROM Tbl_Historial_Resoluciones 
    ORDER BY Fecha_Resolucion DESC
    """
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)
    return sanitize(df).to_dict(orient='records')

def log_update(message):
    try:
        base_dir = get_base_path()
        log_path = os.path.join(base_dir, "debug_sql_log.txt")
        with open(log_path, "a", encoding="utf-8") as f:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{timestamp}] {message}\n")
    except:
        pass

def update_master(code, payload, force_resolve=False, resolution_status=None):
    log_update(f"Attempting UPDATE for {code}. Force: {force_resolve}, Status: {resolution_status}")
    engine = get_engine()
    
    status_resolution = 'IGNORADO'
    if resolution_status:
        status_resolution = resolution_status
    elif force_resolve:
         if payload and payload.get('Descripcion'):
             status_resolution = 'CORREGIDO'

    history_logged = False
    
    try:
        with engine.begin() as conn:
            if payload:
                q = text("""
                    UPDATE Tbl_Maestro_Piezas
                    SET Descripcion=:d, Material=:m, Medida=:md,
                        Proceso_Primario=:p0, Proceso_1=:p1, Proceso_2=:p2, Proceso_3=:p3,
                        Ultima_Actualizacion=GETDATE()
                    WHERE Codigo_Pieza=:c
                """)
                conn.execute(q, {
                    'c': code,
                    'd': payload.get('Descripcion'),
                    'm': payload.get('Material'),
                    'md': payload.get('Medida'),
                    'p0': payload.get('Proceso_Primario'),
                    'p1': payload.get('Proceso_1'),
                    'p2': payload.get('Proceso_2'),
                    'p3': payload.get('Proceso_3')
                })

            conn.execute(text("UPDATE Tbl_Historial_Proyectos SET Requiere_Correccion = 0, Estado_Resolucion = :s WHERE Codigo_Pieza = :c"), {'c': code, 's': status_resolution})

            if force_resolve:
                qh = text("""
                    INSERT INTO Tbl_Historial_Resoluciones 
                    (Codigo_Pieza, Descripcion_Final, Estado_Resolucion, Fecha_Resolucion, Usuario)
                    VALUES (:c, :d, :st, GETDATE(), 'SISTEMA')
                """)
                
                desc_final = payload.get('Descripcion') if payload else 'VALOR ORIGINAL CONSERVADO'
                
                conn.execute(qh, {
                    'c': code,
                    'd': desc_final,
                    'st': status_resolution
                })
                history_logged = True

        return {"status": "success", "history_logged": history_logged}
    except Exception as e:
        log_update(f"Error in update_master: {e}")
        return {"status": "error", "message": str(e), "history_logged": False}

def insert_master(payload):
    engine = get_engine()
    q = text("""
        INSERT INTO Tbl_Maestro_Piezas (Codigo_Pieza, Descripcion, Material, Medida, Proceso_Primario, Proceso_1, Proceso_2, Proceso_3)
        VALUES (:c, :d, :m, :md, :p0, :p1, :p2, :p3)
    """)
    with engine.begin() as conn:
        conn.execute(q, {
            'c': payload.get('Codigo_Pieza'), 
            'd': payload.get('Descripcion'), 
            'm': payload.get('Material'), 
            'md': payload.get('Medida'),
            'p0': payload.get('Proceso_Primario'),
            'p1': payload.get('Proceso_1'), 
            'p2': payload.get('Proceso_2'), 
            'p3': payload.get('Proceso_3')
        })
def delete_master(code):
    engine = get_engine()
    q = text("DELETE FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = :c")
    with engine.begin() as conn:
        conn.execute(q, {'c': code})
    return {"status": "success"}

def fetch_part(code):
    engine = get_engine()
    q = text("""
        SELECT TOP 1 
            Codigo_Pieza, Descripcion, 
            ISNULL(Material, '') as Material, 
            ISNULL(Medida, '') as Medida,
            ISNULL(Proceso_Primario, '') as Proceso_Primario,
            ISNULL(Proceso_1, '') as Proceso_1,
            ISNULL(Proceso_2, '') as Proceso_2,
            ISNULL(Proceso_3, '') as Proceso_3
        FROM Tbl_Maestro_Piezas WHERE Codigo_Pieza = :code
    """)
    with engine.connect() as conn:
        df = pd.read_sql(q, conn, params={"code": code})
    return sanitize(df).to_dict(orient='records')

def get_homologation(code):
    engine = get_engine()
    query = text("SELECT * FROM V_Auditoria_Conflictos WHERE Codigo_Pieza = :c")
    with engine.connect() as conn:
        df = pd.read_sql(query, conn, params={"c": code})
    return sanitize(df).to_dict(orient='records')


def get_pending_tasks():
    engine = get_engine()
    # Usar VISTA DE CONFLICTOS como fuente principal (684 registros detectados)
    query = "SELECT * FROM V_Auditoria_Conflictos"
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)
    
    # SOLUCIÓN MAESTRA v10.7: DATA TRANSLATOR logic
    
    # 1. Normalizar nombres de columnas (Todo a minúsculas para evitar case-sensitivity)
    df.columns = [c.lower() for c in df.columns]

    # 2. Renombrar columnas específicas (SQL -> Flutter Key)
    rename_map = {
        'id': 'id', # Asegurar ID si existe
        'archivo_origen': 'archivo',
        'nombre_archivo': 'archivo',
        'file': 'archivo',
        'nombre_hoja': 'hoja',
        'sheet': 'hoja',
        'fila_excel': 'fila',
        'numero_fila_excel': 'fila', # Added for safety
        'row': 'fila',
        'codigo_pieza': 'codigo',
        'parte': 'codigo',
        'desc_excel': 'desc_excel',
        'descripcion_excel': 'desc_excel'
    }
    
    # Búsqueda de Candidatos para Descripción (Detectivazo)
    candidatos_desc = ['descripcion_base', 'desc_master', 'descripcion', 'desc_oficial', 'descripcion_final', 'desc']
    col_encontrada = next((c for c in candidatos_desc if c in df.columns), None)
    
    if col_encontrada:
        rename_map[col_encontrada] = 'descripcion'
    
    df = df.rename(columns=rename_map)
    
    # Si de verdad no existe, llenar con un texto de aviso técnico
    if 'descripcion' not in df.columns:
        df['descripcion'] = 'Columna Descripción No Encontrada en SQL'

    # 3. Sanitizar Nulos (Blindaje)
    if not df.empty:
        # Usar .get para evitar KeyErrors si la columna no existe tras el rename
        if 'archivo' not in df.columns: df['archivo'] = 'Desconocido.xlsx'
        else: df['archivo'] = df['archivo'].fillna('Desconocido.xlsx')

        if 'fila' not in df.columns: df['fila'] = 0
        else: df['fila'] = df['fila'].fillna(0)
        
        if 'hoja' not in df.columns: df['hoja'] = 'Sheet1'
        else: df['hoja'] = df['hoja'].fillna('Sheet1')

    return sanitize(df).to_dict(orient='records')

def export_master():
    try:
        engine = get_engine()
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        filename = f"Maestro_Materiales_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(desktop, filename)
        
        query = "SELECT * FROM Tbl_Maestro_Piezas ORDER BY Codigo_Pieza"
        with engine.connect() as conn:
            df = pd.read_sql(query, conn)
            df.to_excel(output_path, index=False)
            
        return {"status": "success", "path": output_path}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def mark_task_solved(id):
    engine = get_engine()
    q = text("UPDATE Tbl_Historial_Proyectos SET Requiere_Correccion = 0 WHERE Id = :id")
    with engine.begin() as conn:
        conn.execute(q, {"id": id})
    return {"status": "success"}

def register_file_path(filename, full_path):
    pmap = load_path_map()
    pmap[filename] = full_path
    save_path_map(pmap)
    return {"status": "success", "message": f"Ruta actualizada para {filename}"}

def write_excel_correction(id, new_value, filename, sheet_name, row_idx, col_name):
    # 1. Resolver Ruta
    pmap = load_path_map()
    full_path = pmap.get(filename)
    
    if not full_path or not os.path.exists(full_path):
        return {"status": "error", "message": f"Ruta no encontrada para '{filename}'. Vaya a 'Fuentes de Datos' y relocalice el archivo."}
        
    try:
        # 2. Abrir Excel
        wb = openpyxl.load_workbook(full_path)
        
        if sheet_name not in wb.sheetnames:
             return {"status": "error", "message": f"Hoja '{sheet_name}' no existe en {filename}"}
             
        ws = wb[sheet_name]
        
        # 3. Calcular Coordenada
        # Asumimos que col_name es una letra (ej: 'C') o buscamos por encabezado si fuera necesario.
        # En v13.1, asumiremos que el frontend o la configuración nos dice qué columna es 'Descripcion'. 
        # Si no, por defecto intentaremos buscar la columna 'Descripcion' en la fila 1.
        
        target_col_idx = None
        
        # Estrategia de búsqueda de columna (Simplificada para v13.1)
        # Buscamos en la fila 1 headers como 'Descripcion', 'Desc', 'Description'
        for cell in ws[1]:
            if cell.value and str(cell.value).lower() in ['descripcion', 'descripción', 'desc', 'description']:
                target_col_idx = cell.column
                break
        
        if not target_col_idx:
             # Fallback: Usar columna C (3) como estándar si no se encuentra header
             target_col_idx = 3 
             
        # Fila: openpyxl es 1-based. Si row_idx viene de dataframe (0-based) o SQL, ajustar.
        # Generalmente SQL almacena la fila real de Excel. Asumimos row_idx es el número visual de fila.
        try:
            r = int(row_idx)
        except:
            return {"status": "error", "message": f"Índice de fila inválido: {row_idx}"}

        # 4. Manejo de MERGED CELLS
        target_cell = ws.cell(row=r, column=target_col_idx)
        final_target = target_cell
        
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                # Si está combinada, escribir en la celda superior izquierda del rango
                # bounds devuelve (min_col, min_row, max_col, max_row)
                min_col, min_row, max_col, max_row = merged_range.bounds
                final_target = ws.cell(row=min_row, column=min_col)
                break
        
        # 5. Escribir Valor
        final_target.value = new_value
        
        # 6. Guardar (Manejo de Permisos)
        wb.save(full_path)
        wb.close()
        
        # 7. Actualizar SQL para reflejar que se corrigió en Excel
        # Opcional: Marcar como 'CORREGIDO_EN_EXCEL' o simplemente 'CORREGIDO'
        save_excel_correction(id, new_value) 
        
        return {"status": "success", "message": "Excel actualizado correctamente."}
        
    except PermissionError:
        return {"status": "error", "message": f"El archivo '{filename}' está ABIERTO por otro usuario. Ciérrelo e intente de nuevo."}
    except Exception as e:
        return {"status": "error", "message": f"Error escribiendo Excel: {str(e)}"}

def save_excel_correction(id, new_desc):
    try:
        engine = get_engine()
        with engine.begin() as conn:
            conn.execute(text("UPDATE Tbl_Auditoria_Conflictos SET Estado = 'CORREGIDO', Desc_Excel = :d WHERE ID = :id"), {"d": new_desc, "id": id})
        return {"status": "success"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def find_blueprint(code):
    try:
        cfg = load_config()
        bp_path = cfg.get('blueprints_path', '')
        generics_path = cfg.get('generics_path', r"Z:\5. PIEZAS GENERICAS\JA'S PDF")
        
        code_clean = code.strip().upper()
        
        if code_clean.startswith('JA'):
            direct_path = os.path.join(generics_path, f"{code_clean}.pdf")
            if os.path.exists(direct_path):
                return {"status": "success", "path": direct_path, "level": "static_ja"}
        
        if '-' in code_clean:
            prefix = code_clean.split('-')[0]
            direct_guess = os.path.join(bp_path, prefix, f"{code_clean}.pdf")
            if os.path.exists(direct_guess):
                return {"status": "success", "path": direct_guess, "level": "direct_optimization"}
            
        if not bp_path or not os.path.exists(bp_path):
            return {"status": "error", "message": "Ruta de planos no configurada"}
            
        matches = []
        SKIP_DIRS = ['OBSOLETO', 'RESPALDO', 'BACKUP', 'OLD', 'BAK']
        
        search_root = bp_path
        if '-' in code_clean:
            prefix = code_clean.split('-')[0]
            try:
                potential_dirs = [d for d in os.listdir(bp_path) if os.path.isdir(os.path.join(bp_path, d))]
                for d in potential_dirs:
                    if prefix in d.upper():
                        search_root = os.path.join(bp_path, d)
                        break
            except:
                pass

        for root, dirs, files in os.walk(search_root):
            dirs[:] = [d for d in dirs if not any(s in d.upper() for s in SKIP_DIRS)]
            if f"{code_clean}.pdf" in files:
                full_path = os.path.join(root, f"{code_clean}.pdf")
                return {"status": "success", "path": full_path, "level": "deep_search"}
        
        return {"status": "error", "message": "Archivo no encontrado"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def run_full_diagnostics():
    # Estructura plana solicitada por el usuario v10.5
    response = {
        "db_status": False,
        "integrity_status": False,
        "logic_status": False,
        "path_status": False,
        "log": ""
    }
    
    log_buffer = []

    def add_log(msg):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        log_buffer.append(f"[{timestamp}] {msg}")

    try:
        add_log("=== INICIANDO DIAGNÓSTICO SENTINEL PRO v10.5 ===")
        
        # PASO 1: Conexión SQL
        try:
            add_log("1. Verificando Conexión SQL...")
            engine = get_engine()
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            response['db_status'] = True
            add_log("✅ Conexión SQL: EXITOSA")
        except Exception as e:
            add_log(f"❌ Error Crítico de Conexión: {str(e)}")
            raise e # Detener todo si no hay DB

        # PASO 2: Auditoría de Tablas
        try:
            add_log("2. Auditando Esquema de Tablas...")
            with engine.connect() as conn:
                tables = ['Tbl_Maestro_Piezas', 'Tbl_Historial_Resoluciones', 'Tbl_Historial_Proyectos']
                missing = []
                for t in tables:
                    try:
                        conn.execute(text(f"SELECT TOP 1 * FROM {t}"))
                        add_log(f"  - Tabla '{t}': OK")
                    except:
                        add_log(f"  - Tabla '{t}': NO ENCONTRADA")
                        missing.append(t)
                
                if missing:
                    add_log("❌ Tablas faltantes detectadas.")
                else:
                    try:
                        conn.execute(text("SELECT TOP 1 Nombre_Archivo, Nombre_Hoja, Numero_Fila_Excel FROM Tbl_Historial_Proyectos"))
                        add_log("  - Columnas de Metadatos Excel: OK")
                        response['integrity_status'] = True
                    except Exception as ce:
                        add_log(f"❌ Faltan columnas críticas en Tbl_Historial_Proyectos: {ce}")
        except Exception as e:
            add_log(f"❌ Error en auditoría de tablas: {e}")

        # PASO 3: Simulación Lógica Excel
        try:
            add_log("3. Simulando Lógica de Conflictos (Dry Run)...")
            query = "SELECT * FROM V_Auditoria_Conflictos"
            with engine.connect() as conn:
                df = pd.read_sql(query, conn)
            
            count = len(df)
            add_log(f"  - Filas leídas de V_Auditoria_Conflictos: {count}")
            
            if count == 0:
                add_log("⚠️ CERO conflictos detectados. ¿Lectura de Excel vacía?")
                # Warning status is technically "True" for logic operation but with warning log
                response['logic_status'] = True 
            else:
                add_log("✅ Lógica de Vistas operando correctamente.")
                response['logic_status'] = True

        except Exception as e:
            add_log(f"❌ Error leyendo vista de conflictos: {e}")

        # PASO 4: Integridad de Rutas
        try:
            add_log("4. Verificando Acceso a Recursos...")
            cfg = load_config()
            bp_path = cfg.get('blueprints_path', '')
            if not bp_path:
                 add_log("⚠️ Ruta de planos no configurada.")
            elif os.path.exists(bp_path):
                 add_log(f"✅ Ruta de Planos accesible: {bp_path}")
                 response['path_status'] = True
            else:
                 add_log(f"❌ Ruta de Planos INACCESIBLE: {bp_path}")
        except Exception as e:
            add_log(f"❌ Error validando rutas: {e}")

    except Exception as general_error:
        add_log(f"\n❌ EXCEPCIÓN GENERAL DEL SISTEMA: {str(general_error)}")
    
    finally:
        # Construir reporte final
        full_log = "\n".join(log_buffer)
        response['log'] = full_log
        
        # Guardar reporte físico en el Escritorio (USERPROFILE/Desktop)
        try:
            desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
            report_path = os.path.join(desktop, "SENTINEL_LOG_v10.txt")
            with open(report_path, "w", encoding="utf-8") as f:
                f.write(full_log)
        except:
            pass # No fallar si no se puede escribir el archivo

        return response

# --- EXECUTION ---
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('command', nargs='?', help='API Command') # Optional for loop mode
    parser.add_argument('--code', help='Part code')
    parser.add_argument('--id', help='Task ID')
    parser.add_argument('--stdin', action='store_true', help='Read payload from stdin (base64)')
    parser.add_argument('--force_resolve', action='store_true', help='Force resolution')
    parser.add_argument('--status', help='Custom resolution status')
    parser.add_argument('--listen', action='store_true', help='Start in persistent listener mode')
    
    args = parser.parse_known_args()[0]

    def process_command(cmd, payload, args_obj):
        try:
            result = None
            if cmd == 'test_connection':
                result = test_connection()
            elif cmd in ['get_all', 'catalog']:
                result = get_master_catalog()
            elif cmd in ['conflicts', 'get_conflicts']:
                result = get_conflicts()
            elif cmd in ['history', 'get_history']:
                code_val = args_obj.code if args_obj else payload.get('code')
                result = get_history(code_val)
            elif cmd == 'update':
                code_val = args_obj.code if args_obj else payload.get('code')
                force = args_obj.force_resolve if args_obj else payload.get('force_resolve')
                status = args_obj.status if args_obj else payload.get('status')
                result = update_master(code_val, payload, force, status)
            elif cmd == 'delete':
                code_val = args_obj.code if args_obj else payload.get('code')
                result = delete_master(code_val)
            elif cmd == 'insert':
                result = insert_master(payload)
            elif cmd in ['fetch', 'fetch_part']:
                code_val = args_obj.code if args_obj else payload.get('code')
                result = fetch_part(code_val)
            elif cmd in ['homologation', 'get_homologation']:
                code_val = args_obj.code if args_obj else payload.get('code')
                result = get_homologation(code_val)
            elif cmd == 'get_resolved':
                result = get_resolved_tasks()
            elif cmd == 'get_pending':
                result = get_pending_tasks()
            elif cmd in ['mark_corrected', 'mark_solved']:
                id_val = args_obj.id if args_obj and args_obj.id else (args_obj.code if args_obj else payload.get('id'))
                result = mark_task_solved(id_val)
            elif cmd == 'find_blueprint':
                code_val = args_obj.code if args_obj else payload.get('code')
                result = find_blueprint(code_val)
            elif cmd == 'export_master':
                result = export_master()
            elif cmd == 'diagnostic':
                result = run_full_diagnostics()
            # --- COMMANDS v12.0 STANDARDS ---
            elif cmd in ['standards', 'get_standards']:
                result = get_standards()
            elif cmd == 'add_standard':
                desc = payload.get('Descripcion') if payload else (args_obj.code if args_obj else '')
                cat = payload.get('Categoria', 'GENERAL') if payload else 'GENERAL'
                result = add_standard(desc, cat)
            elif cmd == 'edit_standard':
                id_val = args_obj.id if args_obj else payload.get('id')
                new_desc = payload.get('Descripcion') if payload else (args_obj.code if args_obj else '')
                result = edit_standard(id_val, new_desc)
            elif cmd == 'delete_standard':
                id_val = args_obj.id if args_obj else payload.get('id')
                result = delete_standard(id_val)
            # --- COMMANDS v12.1 SMART HOMOLOGATOR ---
            elif cmd == 'get_suggestion':
                dirty = (args_obj.code if args_obj else None) or (payload.get('text') if payload else None)
                result = get_match_suggestion(dirty)
            elif cmd == 'save_correction':
                id_val = args_obj.id if args_obj else payload.get('id')
                txt = payload.get('text')
                result = save_excel_correction(id_val, txt)
            # --- COMMANDS v13.x DATA MANAGER & CONFIG ---
            elif cmd == 'get_config':
                result = load_config()
            elif cmd == 'save_config':
                result = save_sys_config(payload)
            elif cmd in ['get_sources', 'get_paths']:
                result = get_sources()
            elif cmd in ['add_source', 'register_path']:
                name = payload.get('name') or payload.get('filename') or "Archivo Nuevo"
                path = payload.get('path')
                result = add_source(name, path)
            elif cmd == 'update_source':
                result = update_source(payload.get('id'), payload.get('path'))
            elif cmd == 'scan_source':
                result = scan_and_ingest(payload.get('id'))
            elif cmd == 'write_excel':
                result = write_excel_correction(
                    payload.get('id'), 
                    payload.get('value'), 
                    payload.get('filename'), 
                    payload.get('sheet'), 
                    payload.get('row'),
                    'D' 
                )
            elif cmd == 'kill':
                sys.exit(0)
            else:
                result = {"status": "error", "message": f"Comando desconocido: {cmd}"}
                
            return result
        except Exception as e:
            return {"status": "error", "message": str(e)}

    # --- MODE 1: PERSISTENT LISTENER (OPTIMIZATION v14.1) ---
    if args.listen:
        # Optimización: Mantener proceso vivo para evitar carga repetitiva de Python/Librerías
        while True:
            try:
                # 1. FRENO DE MANO: Pausa obligatoria
                time.sleep(0.05) # 50ms es suficiente para respuesta rápida sin quemar CPU

                # 2. DETECCIÓN DE PADRE MUERTO (Suicide Protocol)
                if sys.stdin.closed:
                    sys.exit(0)

                # Leer línea de stdin
                line = sys.stdin.readline()
                if not line:
                    # EOF recibido (padre cerró el stream)
                    sys.exit(0)
                
                line = line.strip()
                if not line:
                    continue

                # Procesar Request
                try:
                    req = json.loads(line)
                    cmd = req.get('command')
                    payload = req.get('payload', {})
                    
                    # Logica de args (shim para compatibilidad con funcion process_command)
                    # En modo listen, todo viene en payload
                    
                    result = process_command(cmd, payload, None)
                    
                    # Responder
                    print(json.dumps(result, default=str))
                    sys.stdout.flush() # CRITICO: Enviar inmediatamente

                except json.JSONDecodeError:
                    print(json.dumps({"status": "error", "message": "JSON invalido"}, default=str))
                    sys.stdout.flush()
                    
            except KeyboardInterrupt:
                sys.exit(0)
            except Exception as e:
                # 3. LOGGING CONTROLADO Y WAIT
                # Si falla el bucle (ej. stdin roto), esperar antes de reintentar o morir
                try:
                    print(json.dumps({"status": "error", "message": f"Loop Error: {str(e)}"}, default=str))
                    sys.stdout.flush()
                except:
                    pass
                time.sleep(1.0) # Espera larga si hay error grave

    # --- MODE 2: ONE-SHOT (LEGACY) ---
    else:
        payload = None
        if args.stdin:
            try:
                stdin_data = sys.stdin.read().strip()
                if stdin_data:
                    try:
                        payload = json.loads(base64.b64decode(stdin_data).decode('utf-8'))
                    except:
                        payload = json.loads(stdin_data)
            except:
                pass
        
        # Shim para process_command con CLI args
        # Convertimos args a objeto compatible o usamos payload
        # process_command usa args_obj para .code, .id, etc.
        res = process_command(args.command, payload, args)
        print(json.dumps(res, default=str))

# --- ESTÁNDARES DE MATERIALES (v12.0) ---

DEFAULT_STANDARDS = [
    {"Descripcion": "ACERO ASTM A36 1/8\"", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A36 3/16\"", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A36 C.10", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A36 C.14", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A36 C.16", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A572 G50 1/2\"", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A572 G50 1/4\"", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A572 G50 3/4\"", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A572 G50 3/8\"", "Categoria": "ACERO"},
    {"Descripcion": "ACERO ASTM A572 G50 5/16\"", "Categoria": "ACERO"},
    {"Descripcion": "ACERO INOXIDABLE 304 CAL.16", "Categoria": "ACERO INOXIDABLE"},
    {"Descripcion": "ACERO INOXIDABLE C.11", "Categoria": "ACERO INOXIDABLE"},
    {"Descripcion": "ALUMINIO 3003 C.11", "Categoria": "ALUMINIO"},
    {"Descripcion": "ALUMINIO 3003 C.14", "Categoria": "ALUMINIO"},
    {"Descripcion": "ALUMINIO 5052 1/4\"", "Categoria": "ALUMINIO"},
    {"Descripcion": "ALUMINIO NEGRO 3003 C.19", "Categoria": "ALUMINIO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 1 1/2\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 1/2\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 2 1/2\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 2\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 3 1/2\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 3\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 4\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ALUMINIO MACIZO 6026 Ø 7/8\"", "Categoria": "ALUMINIO MACIZO"},
    {"Descripcion": "ANGULO ASTM A36 1 1/2\" x 1 1/2\" x 3/16\"", "Categoria": "ANGULO ASTM A36"},
    {"Descripcion": "ANGULO ASTM A36 1\" x 1\" x 3/16\"", "Categoria": "ANGULO ASTM A36"},
    {"Descripcion": "ANGULO ASTM A36 2\" x 2\" x 3/16\"", "Categoria": "ANGULO ASTM A36"},
    {"Descripcion": "BARRA HUECA AISI 1018 Ø 33mm x 14mm", "Categoria": "BARRA HUECA AISI 1018"},
    {"Descripcion": "BARRA HUECA AISI 1018 Ø 40mm x 25mm", "Categoria": "BARRA HUECA AISI 1018"},
    {"Descripcion": "BARRA HUECA AISI 1018 Ø 40mm x 28mm", "Categoria": "BARRA HUECA AISI 1018"},
    {"Descripcion": "BARRA HUECA AISI 1018 Ø 50mm x 35mm", "Categoria": "BARRA HUECA AISI 1018"},
    {"Descripcion": "BARRA HUECA AISI 1018 Ø 76mm x 38mm", "Categoria": "BARRA HUECA AISI 1018"},
    {"Descripcion": "BARRA CROMADA AISI 1045 Ø 28mm", "Categoria": "BARRA CROMADA AISI 1045"},
    {"Descripcion": "BARRA CROMADA AISI 1045 Ø 45mm", "Categoria": "BARRA CROMADA AISI 1045"},
    {"Descripcion": "TUBO HONEADO AISI 1018 Ø 60mm x 50mm", "Categoria": "TUBO HONEADO AISI 1018"},
    {"Descripcion": "TUBO HONEADO AISI 1018 Ø 73mm x 63mm", "Categoria": "TUBO HONEADO AISI 1018"},
    {"Descripcion": "BARRA HUECA CROMADA AISI 1018 Ø 38.1mm X 25.4mm", "Categoria": "BARRA HUECA CROMADA"},
    {"Descripcion": "BARRA HUECA DE ALUMINIO B241 6026 Ø 101.6mm X 50.5mm", "Categoria": "BARRA HUECA DE ALUMINIO"},
    {"Descripcion": "BARRA HUECA DE ALUMINIO B241 6026 Ø 63.5mm X 29.7mm", "Categoria": "BARRA HUECA DE ALUMINIO"},
    {"Descripcion": "BARRA HUECA DE ALUMINIO B241 6026 Ø 76.2mm X 29.7mm", "Categoria": "BARRA HUECA DE ALUMINIO"},
    {"Descripcion": "BARRA HUECA DE ALUMINIO B241 6026 Ø 88.9mm X 24.7mm", "Categoria": "BARRA HUECA DE ALUMINIO"},
    {"Descripcion": "CAJA DE TENSADO DE LONA", "Categoria": "ACCESORIOS"},
    {"Descripcion": "CANAL C A36 4\"", "Categoria": "PERFIL"},
    {"Descripcion": "COMERCIAL BISAGRA DE LIBRO", "Categoria": "COMERCIAL"},
    {"Descripcion": "COMERCIAL BISAGRA DE PIANO", "Categoria": "COMERCIAL"},
    {"Descripcion": "MATRACA DE LONA", "Categoria": "ACCESORIOS"},
    {"Descripcion": "PERFIL ALUMINIO PELDAÑO 688 6061 T6", "Categoria": "ALUMINIO"},
    {"Descripcion": "PERNO REY COMERCIAL", "Categoria": "COMERCIAL"},
    {"Descripcion": "SEGURO DE FUNDICION -", "Categoria": "COMERCIAL"},
    {"Descripcion": "SEGURO DE RESORTE CORTO", "Categoria": "COMERCIAL"},
    {"Descripcion": "SEGURO DE RESORTE LARGO", "Categoria": "COMERCIAL"},
    {"Descripcion": "HSS ASTM A500 °B 2 1/2\" x 2 1/2\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 2 1/2\" x 2 1/2\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 2\" x 2\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 2\" x 2\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 3 1/2\" X 3 1/2\" X 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 3\" x 2\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 3\" x 2\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 3\" x 3\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 3\" x 3\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 4 1/2\" x 3 1/2\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 4\" x 2\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 4\" x 2\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 4\" x 3\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 4\" x 3\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 4\" x 3\" x 3/8\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 4\" x 4\" x 3/8\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" x 2\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" x 2\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" x 3\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" x 3\" x 3/16\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" x 4\" x 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" x 4\" x 3/8\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" X 6\" X 1/4\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "HSS ASTM A500 °B 6\" x 6\" x 3/8\"", "Categoria": "HSS ASTM A500"},
    {"Descripcion": "PLACA HARDOX 1/4\"", "Categoria": "PLACA"},
    {"Descripcion": "PLACA STRENX 110 XF 3/16\"", "Categoria": "PLACA"},
    {"Descripcion": "PLACA STRENX 110XF 1/2\"", "Categoria": "PLACA"},
    {"Descripcion": "PTR ASTM A36 1 1/2\" x 1 1/2 \" x 3/16\"", "Categoria": "PTR ASTM A36"},
    {"Descripcion": "PTR ASTM A36 1\" x 1\" x C.11", "Categoria": "PTR ASTM A36"},
    {"Descripcion": "REDONDO AISI 1018 Ø 1 1/2\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 1 1/4\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 1 3/8\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 1\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 1/2\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 2 1/2\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 2 5/8\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 2\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 3\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 3/4\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO AISI 1018 Ø 7/8\"", "Categoria": "REDONDO AISI 1018"},
    {"Descripcion": "REDONDO NEGRO Ø 5/16\"", "Categoria": "REDONDO NEGRO"},
    {"Descripcion": "REDONDO NEGRO Ø 5/8\"", "Categoria": "REDONDO NEGRO"},
    {"Descripcion": "RIEL DE ACERO A36 1500", "Categoria": "PERFIL"},
    {"Descripcion": "SOLERA ASTM A36 1 1/2\" x 1/2\"", "Categoria": "SOLERA ASTM A36"},
    {"Descripcion": "SOLERA ASTM A36 1 1/4\" x 1/4\"", "Categoria": "SOLERA ASTM A36"},
    {"Descripcion": "SOLERA ASTM A36 1\" x 1/2\"", "Categoria": "SOLERA ASTM A36"},
    {"Descripcion": "SOLERA ASTM A36 2\" x 1\"", "Categoria": "SOLERA ASTM A36"},
    {"Descripcion": "SOLERA ASTM A36 4\" x 1\"", "Categoria": "SOLERA ASTM A36"},
    {"Descripcion": "SOLERA ASTM A36 6\" x 1\"", "Categoria": "SOLERA ASTM A36"},
    {"Descripcion": "SOLERA DE ALUMINIO ASTM A36 2\" X 1\"", "Categoria": "SOLERA DE ALUMINIO"},
    {"Descripcion": "TOLDO ALUMINIO C.19", "Categoria": "TOLDO ALUMINIO"},
    {"Descripcion": "TUBO DE ACERO A500 °B Ø 1 1/2\" CED. 80", "Categoria": "TUBO DE ACERO A500"},
    {"Descripcion": "TUBO DE ACERO A500 °B Ø 1 1/2\" CED. 80 SIN/COS", "Categoria": "TUBO DE ACERO A500"},
    {"Descripcion": "TUBO DE ACERO A500 °B Ø 1\" CED. 40 C/COS", "Categoria": "TUBO DE ACERO A500"},
    {"Descripcion": "TUBO DE ACERO A500 °B Ø 1\" CED. 40 SIN/COS", "Categoria": "TUBO DE ACERO A500"},
    {"Descripcion": "TUBO DE ALUMINIO B241 Ø  2\" x  1\" x 1/8\"", "Categoria": "TUBO DE ALUMINIO B241"},
    {"Descripcion": "TUBO DE ALUMINIO B241 Ø 2 1/2\"", "Categoria": "TUBO DE ALUMINIO B241"},
    {"Descripcion": "TUBO DE ALUMINIO B241 Ø 2\"", "Categoria": "TUBO DE ALUMINIO B241"},
    {"Descripcion": "TUBO DE ALUMINIO B241 Ø 2\" x  1\" x 1/8\"", "Categoria": "TUBO DE ALUMINIO B241"},
    {"Descripcion": "TUBO DE ALUMINIO B241 Ø 3 1/2\"", "Categoria": "TUBO DE ALUMINIO B241"},
    {"Descripcion": "TUBO DE ALUMINIO NEGRO B241 Ø 2 1/2\"", "Categoria": "TUBO DE ALUMINIO B241"},
    {"Descripcion": "TUBO STROCK CROMADO ASTM 1045 Ø 70mm X 63mm", "Categoria": "TUBO STROCK CROMADO"},
    {"Descripcion": "PERFIL ALUMINIO CUERNO EA 685 6061T6", "Categoria": "PERFIL ALUMINIO"},
    {"Descripcion": "PERFIL ALUMINIO PRINCIPAL EXT 684 6061T6", "Categoria": "PERFIL ALUMINIO"},
    {"Descripcion": "PERFIL ALUMINIO ANGULO VISTA 686", "Categoria": "PERFIL ALUMINIO"},
    {"Descripcion": "PERFIL ALUMINIO REFUERZO INT 683", "Categoria": "PERFIL ALUMINIO"},
    {"Descripcion": "BORDA LATERAL BASCULANTE 4.9 6061", "Categoria": "BORDA LATERAL"},
    {"Descripcion": "BORDA LATERAL BASCULANTE 3.5 6061", "Categoria": "BORDA LATERAL"},
    {"Descripcion": "PERFIL DE ALUMINIO TIPO BISAGRA ABATIBLE 3.10 MT 6061-T6", "Categoria": "ALUMINIO"},
    {"Descripcion": "PERFIL DE ALUMINIO TIPO ESCALON ABATIBLE 3.10 MT 6061-T6", "Categoria": "ALUMINIO"}
]

def ensure_standards_table():
    engine = get_engine()
    
    # 1. Crear tabla si no existe
    create_table_query = """
    IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='Tbl_Estandares_Materiales' AND xtype='U')
    BEGIN
        CREATE TABLE Tbl_Estandares_Materiales (
            ID INT IDENTITY(1,1) PRIMARY KEY, 
            Descripcion NVARCHAR(400) UNIQUE NOT NULL, 
            Categoria NVARCHAR(100)
        )
    END
    """
    
    try:
        with engine.begin() as conn:
            conn.execute(text(create_table_query))
            
            # 2. Verificar si está vacía para sembrar datos
            res = conn.execute(text("SELECT COUNT(*) FROM Tbl_Estandares_Materiales")).fetchone()
            count = res[0] if res else 0
            
            if count == 0:
                print("Sembrando Tbl_Estandares_Materiales con datos por defecto...")
                insert_query = text("INSERT INTO Tbl_Estandares_Materiales (Descripcion, Categoria) VALUES (:d, :c)")
                for item in DEFAULT_STANDARDS:
                    try:
                        conn.execute(insert_query, {"d": item["Descripcion"], "c": item["Categoria"]})
                    except:
                        pass # Ignorar duplicados si por alguna razón fallara la lógica de conteo
    except Exception as e:
        print(f"Error inicializando tabla de estándares: {e}")

def get_standards():
    ensure_standards_table() # Asegurar existencia antes de leer
    engine = get_engine()
    with engine.connect() as conn:
        df = pd.read_sql("SELECT * FROM Tbl_Estandares_Materiales ORDER BY Descripcion ASC", conn)
    return sanitize(df).to_dict(orient='records')

def add_standard(desc, cat="GENERAL"):
    engine = get_engine()
    try:
        with engine.begin() as conn:
            conn.execute(text("INSERT INTO Tbl_Estandares_Materiales (Descripcion, Categoria) VALUES (:d, :c)"), {"d": desc, "c": cat})
        return {"status": "success"}
    except Exception as e:
        if "UNIQUE constraint" in str(e) or "2627" in str(e):
            return {"status": "error", "message": "El material ya existe en la biblioteca."}
        return {"status": "error", "message": str(e)}

def edit_standard(id, new_desc):
    engine = get_engine()
    try:
        with engine.begin() as conn:
            conn.execute(text("UPDATE Tbl_Estandares_Materiales SET Descripcion = :d WHERE ID = :id"), {"d": new_desc, "id": id})
        return {"status": "success"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def delete_standard(id):
    engine = get_engine()
    try:
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM Tbl_Estandares_Materiales WHERE ID = :id"), {"id": id})
        return {"status": "success"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# --- SMART HOMOLOGATOR (v12.1) ---

_STANDARDS_CACHE = None

def get_match_suggestion(dirty_text):
    global _STANDARDS_CACHE
    if not dirty_text:
        return None
        
    try:
        # 1. Cargar estándares si no están en cache (Optimización sugerida)
        if _STANDARDS_CACHE is None:
            engine = get_engine()
            with engine.connect() as conn:
                df = pd.read_sql("SELECT Descripcion FROM Tbl_Estandares_Materiales", conn)
                _STANDARDS_CACHE = df['Descripcion'].tolist()
        
        if not _STANDARDS_CACHE:
            return None
            
        best_match = None
        highest_ratio = 0.0
        
        dirty_clean = dirty_text.strip().upper()
        
        # 2. Lógica Fuzzy con SequenceMatcher
        for standard in _STANDARDS_CACHE:
            ratio = difflib.SequenceMatcher(None, dirty_clean, standard.upper()).ratio()
            if ratio > highest_ratio:
                highest_ratio = ratio
                best_match = standard
                
        # 3. Retornar solo si supera el 60% (Umbral CRÍTICO)
        if highest_ratio >= 0.60:
            return {"suggestion": best_match, "ratio": round(highest_ratio, 2)}
        else:
            return None
            
    except Exception as e:
        return {"status": "error", "message": str(e)}

# --- EXECUTION ---
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('command', help='API Command')
    parser.add_argument('--code', help='Part code')
    parser.add_argument('--id', help='Task ID')
    parser.add_argument('--stdin', action='store_true', help='Read payload from stdin (base64)')
    parser.add_argument('--force_resolve', action='store_true', help='Force resolution')
    parser.add_argument('--status', help='Custom resolution status')
    
    args = parser.parse_known_args()[0]
    
    payload = None
    if args.stdin:
        try:
            stdin_data = sys.stdin.read().strip()
            if stdin_data:
                try:
                    payload = json.loads(base64.b64decode(stdin_data).decode('utf-8'))
                except:
                    payload = json.loads(stdin_data)
        except:
            pass

    cmd = args.command
    
    try:
        result = None
        if cmd == 'test_connection':
            result = test_connection()
        elif cmd in ['get_all', 'catalog']:
            result = get_master_catalog()
        elif cmd in ['conflicts', 'get_conflicts']:
            result = get_conflicts()
        elif cmd in ['history', 'get_history']:
            result = get_history(args.code)
        elif cmd == 'update':
            result = update_master(args.code, payload, args.force_resolve, args.status)
        elif cmd == 'delete':
            result = delete_master(args.code)
        elif cmd == 'insert':
            result = insert_master(payload)
        elif cmd in ['fetch', 'fetch_part']:
            result = fetch_part(args.code)
        elif cmd in ['homologation', 'get_homologation']:
            result = get_homologation(args.code)
        elif cmd == 'get_resolved':
            result = get_resolved_tasks()
        elif cmd == 'get_pending':
            result = get_pending_tasks()
        elif cmd in ['mark_corrected', 'mark_solved']:
            result = mark_task_solved(args.id or args.code)
        elif cmd == 'find_blueprint':
            result = find_blueprint(args.code)
        elif cmd == 'export_master':
            result = export_master()
        elif cmd == 'diagnostic':
            result = run_full_diagnostics()
        # --- COMMANDS v12.0 STANDARDS ---
        elif cmd in ['standards', 'get_standards']:
            result = get_standards()
        elif cmd == 'add_standard':
            desc = payload.get('Descripcion') if payload else args.code
            cat = payload.get('Categoria', 'GENERAL') if payload else 'GENERAL'
            result = add_standard(desc, cat)
        elif cmd == 'edit_standard':
            new_desc = payload.get('Descripcion') if payload else args.code
            result = edit_standard(args.id, new_desc)
        elif cmd == 'delete_standard':
            result = delete_standard(args.id)
        # --- COMMANDS v12.1 SMART HOMOLOGATOR ---
        elif cmd == 'get_suggestion':
            dirty = args.code or (payload.get('text') if payload else None)
            result = get_match_suggestion(dirty)
        elif cmd == 'save_correction':
            result = save_excel_correction(args.id, payload.get('text'))
        # --- COMMANDS v13.x DATA MANAGER & CONFIG ---
        elif cmd == 'get_config':
            result = load_config()
        elif cmd == 'save_config':
            result = save_sys_config(payload)
        elif cmd in ['get_sources', 'get_paths']:
            result = get_sources()
        elif cmd in ['add_source', 'register_path']:
            # Normalizar nombres de llaves (puede venir como 'name' o 'filename')
            name = payload.get('name') or payload.get('filename') or "Archivo Nuevo"
            path = payload.get('path')
            result = add_source(name, path)
        elif cmd == 'update_source':
            result = update_source(payload.get('id'), payload.get('path'))
        elif cmd == 'scan_source':
            # Alias scan_source a scan_and_ingest
            result = scan_and_ingest(payload.get('id'))
        elif cmd == 'write_excel':
            result = write_excel_correction(
                payload.get('id'), 
                payload.get('value'), 
                payload.get('filename'), 
                payload.get('sheet'), 
                payload.get('row'),
                'D' # Default col if needed
            )
        else:
            result = {"status": "error", "message": f"Comando desconocido: {cmd}"}
            
        print(json.dumps(result, default=str))
    except Exception as e:
        print(json.dumps({"status": "error", "message": str(e)}, default=str))
